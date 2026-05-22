"""
services/excel_service.py
Responsável por toda geração de planilhas Excel do sistema.
Extraído de app.py para separar lógica de negócio das rotas Flask.
"""
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Ordem fixa das lojas nas colunas do Excel
ORDEM_LOJAS = ["Pesqueira", "Recife", "Campina Grande", "Natal", "Maceió"]

# Apelido exibido no cabeçalho (Recife aparece como "Cruza")
APELIDO_LOJA = {"Recife": "Cruza"}

# Estilos reutilizáveis
_BROWN  = PatternFill(start_color="8B4513", end_color="8B4513", fill_type="solid")
_RED    = PatternFill(start_color="B71C1C", end_color="B71C1C", fill_type="solid")
_GRAY   = PatternFill(start_color="444444", end_color="444444", fill_type="solid")
_DARK   = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
_WHITE  = Font(color="FFFFFF", bold=True)
_BOLD   = Font(bold=True)
_CENTER = Alignment(horizontal="center", vertical="center")
_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'),  bottom=Side(style='thin')
)


def _nome_col(loja):
    """Retorna o nome a exibir no cabeçalho — aplica apelidos."""
    return APELIDO_LOJA.get(loja.nome, loja.nome)


def ordenar_lojas(lojas):
    """Ordena uma lista de objetos Loja pela ordem personalizada."""
    return sorted(lojas, key=lambda l: ORDEM_LOJAS.index(l.nome) if l.nome in ORDEM_LOJAS else 999)


def gerar_excel_unico(periodo, laboratorio_nome, lojas_filtradas, db, Produto, ItemPedido, Pedido, Negociacao):
    """
    Gera um Excel com os pedidos de UM laboratório para envio por e-mail.
    Retorna um objeto BytesIO pronto para anexar.

    Colunas: PRODUTO | CX EMB | <lojas> | TOTAL | PREÇO TABELA | DESC % | PREÇO LÍQUIDO | VALOR FECHADO | BONIF %

    Parâmetros:
        periodo          — objeto Periodo do banco
        laboratorio_nome — string com o nome do laboratório
        lojas_filtradas  — lista de objetos Loja já filtrada pelo usuário
        db               — instância do SQLAlchemy
        Produto          — model Produto (injetado para evitar import circular)
        ItemPedido       — model ItemPedido
        Pedido           — model Pedido
        Negociacao       — model Negociacao
    """
    from openpyxl.utils import get_column_letter

    lojas = ordenar_lojas(lojas_filtradas)
    n_lojas = len(lojas)

    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    ws = wb.create_sheet(title=laboratorio_nome[:30])

    # ── Índices de coluna ─────────────────────────────────────────────────
    col_produto  = 1
    col_cxemb    = 2
    col_loja_ini = 3
    col_loja_fim = col_loja_ini + n_lojas - 1
    col_total    = col_loja_fim + 1
    col_preco    = col_total + 1
    col_negoc    = col_preco + 1          # início do bloco NEGOCIAÇÃO
    col_desc     = col_negoc
    col_pliq     = col_negoc + 1
    col_vfech    = col_negoc + 2
    col_bonif    = col_negoc + 3
    col_last     = col_bonif

    # ── Linha 1: cabeçalhos fixos e merge do bloco NEGOCIAÇÃO ────────────
    # PRODUTO
    c = ws.cell(row=1, column=col_produto, value="PRODUTO")
    c.font = _WHITE; c.fill = _BROWN; c.alignment = _CENTER; c.border = _BORDER

    # CX EMB
    c = ws.cell(row=1, column=col_cxemb, value="CX EMB")
    c.font = _WHITE; c.fill = _BROWN; c.alignment = _CENTER; c.border = _BORDER

    # Lojas
    for i, loja in enumerate(lojas):
        c = ws.cell(row=1, column=col_loja_ini + i, value=_nome_col(loja))
        c.font = _WHITE; c.fill = _BROWN; c.alignment = _CENTER; c.border = _BORDER

    # TOTAL
    c = ws.cell(row=1, column=col_total, value="TOTAL")
    c.font = _WHITE; c.fill = _RED; c.alignment = _CENTER; c.border = _BORDER

    # PREÇO TABELA
    c = ws.cell(row=1, column=col_preco, value="PREÇO\nTABELA")
    c.font = _WHITE; c.fill = _GRAY
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = _BORDER

    # NEGOCIAÇÃO (merge das 4 sub-colunas)
    ws.merge_cells(start_row=1, start_column=col_negoc, end_row=1, end_column=col_last)
    c = ws.cell(row=1, column=col_negoc, value="NEGOCIAÇÃO")
    c.font = _WHITE; c.fill = _DARK; c.alignment = _CENTER; c.border = _BORDER

    # ── Linha 2: sub-cabeçalhos do bloco NEGOCIAÇÃO ───────────────────────
    # Linha 2 só existe para as colunas de negociação; demais ficam vazias mas bordadas
    for col_idx in [col_produto, col_cxemb] + list(range(col_loja_ini, col_total + 1)) + [col_preco]:
        c = ws.cell(row=2, column=col_idx, value="")
        c.border = _BORDER

    for label, col_idx in [
        ("DESC %",        col_desc),
        ("PREÇO\nLÍQUIDO", col_pliq),
        ("VALOR\nFECHADO", col_vfech),
        ("BONIF %",       col_bonif),
    ]:
        c = ws.cell(row=2, column=col_idx, value=label)
        c.font = _WHITE; c.fill = _DARK
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _BORDER

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 28

    # ── Busca produtos ────────────────────────────────────────────────────
    produtos = Produto.query.filter_by(
        grupo=periodo.grupo_filtro,
        laboratorio=laboratorio_nome
    ).order_by(Produto.nome).all()

    ids_produtos = [p.id for p in produtos]

    # ── Busca quantidades (1 query) ───────────────────────────────────────
    rows_qtd = db.session.query(
        ItemPedido.produto_id,
        Pedido.loja_id,
        db.func.sum(ItemPedido.quantidade).label('total')
    ).join(Pedido, Pedido.id == ItemPedido.pedido_id)\
     .filter(
         Pedido.periodo_id == periodo.id,
         ItemPedido.produto_id.in_(ids_produtos)
     ).group_by(ItemPedido.produto_id, Pedido.loja_id).all()

    qtd_map = {(r.produto_id, r.loja_id): int(r.total or 0) for r in rows_qtd}

    # ── Busca negociações (1 query) ───────────────────────────────────────
    neg_map = {
        n.produto_id: n
        for n in Negociacao.query.filter_by(periodo_id=periodo.id)
                           .filter(Negociacao.produto_id.in_(ids_produtos)).all()
    }

    # ── Dados (a partir da linha 3) ───────────────────────────────────────
    for row, p in enumerate(produtos, start=3):
        # Produto
        c = ws.cell(row=row, column=col_produto, value=p.nome)
        c.border = _BORDER

        # CX EMB
        c = ws.cell(row=row, column=col_cxemb, value=p.unidade_caixa)
        c.alignment = _CENTER; c.border = _BORDER

        # Quantidades por loja
        total_linha = 0
        for i, loja in enumerate(lojas):
            qtd = qtd_map.get((p.id, loja.id), 0)
            c = ws.cell(row=row, column=col_loja_ini + i, value=qtd)
            c.alignment = _CENTER; c.border = _BORDER
            if qtd == 0:
                c.font = Font(color="CCCCCC")
            total_linha += qtd

        # TOTAL
        c = ws.cell(row=row, column=col_total, value=total_linha)
        c.font = Font(color="FFFFFF", bold=True)
        c.fill = _RED; c.alignment = _CENTER; c.border = _BORDER

        # PREÇO TABELA
        c = ws.cell(row=row, column=col_preco, value=p.preco)
        c.number_format = 'R$ #,##0.00'
        c.fill = _GRAY; c.alignment = _CENTER; c.border = _BORDER
        c.font = Font(color="FFFFFF", bold=True)

        # Negociação
        negoc   = neg_map.get(p.id)
        desc    = float(negoc.desconto    if negoc else 0.0)
        bonif   = float(negoc.bonificacao if negoc else 0.0)
        p_liq   = p.preco * (1 - desc / 100)
        v_fech  = p_liq * total_linha

        # DESC %
        c = ws.cell(row=row, column=col_desc, value=desc)
        c.fill = _DARK; c.alignment = _CENTER; c.border = _BORDER
        c.font = Font(color="4FC3F7", bold=True)   # azul claro

        # PREÇO LÍQUIDO
        c = ws.cell(row=row, column=col_pliq, value=p_liq)
        c.number_format = 'R$ #,##0.00'
        c.fill = _DARK; c.alignment = _CENTER; c.border = _BORDER
        c.font = Font(color="A5D6A7", bold=True)   # verde claro

        # VALOR FECHADO
        c = ws.cell(row=row, column=col_vfech, value=v_fech)
        c.number_format = 'R$ #,##0.00'
        c.fill = _DARK; c.alignment = _CENTER; c.border = _BORDER
        c.font = Font(color="A5D6A7", bold=True)

        # BONIF %
        c = ws.cell(row=row, column=col_bonif, value=bonif)
        c.fill = _DARK; c.alignment = _CENTER; c.border = _BORDER
        c.font = Font(color="FFFFFF", bold=True)

    # ── Larguras ──────────────────────────────────────────────────────────
    ws.column_dimensions['A'].width = 30
    for ci in range(col_cxemb, col_last + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 14

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def gerar_excel_consolidado(periodo, lojas, dados_lab):
    """
    Gera o Excel completo de consolidação (todas as abas / todos os laboratórios).
    Retorna um BytesIO pronto para download.

    Parâmetros:
        periodo    — objeto Periodo
        lojas      — lista ordenada de objetos Loja
        dados_lab  — dict {laboratorio: [lista de itens com qtds, preco, desconto, etc.]}
    """
    lojas = ordenar_lojas(lojas)

    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    for lab_nome, itens in dados_lab.items():
        ws = wb.create_sheet(title=lab_nome[:30])

        # Linha 1 — cabeçalhos principais
        ws.merge_cells('A1:A2')
        _cell(ws, 'A1', "PRODUTO", _BROWN)

        ws.merge_cells('B1:B2')
        _cell(ws, 'B1', "CX EMB", _BROWN)

        col_ini = 3
        col_fim = col_ini + len(lojas) - 1
        ws.merge_cells(start_row=1, start_column=col_ini, end_row=1, end_column=col_fim)
        c = ws.cell(row=1, column=col_ini, value="QUANTIDADES POR LOJA")
        c.fill = _BROWN; c.font = _WHITE; c.alignment = _CENTER; c.border = _BORDER

        col_total = col_fim + 1
        ws.merge_cells(start_row=1, start_column=col_total, end_row=2, end_column=col_total)
        c = ws.cell(row=1, column=col_total, value="TOTAL")
        c.fill = _RED; c.font = _WHITE; c.alignment = _CENTER; c.border = _BORDER

        col_preco = col_total + 1
        ws.merge_cells(start_row=1, start_column=col_preco, end_row=2, end_column=col_preco)
        c = ws.cell(row=1, column=col_preco, value="PREÇO TAB")
        c.fill = _GRAY; c.font = _WHITE; c.alignment = _CENTER; c.border = _BORDER

        col_negoc = col_preco + 1
        ws.merge_cells(start_row=1, start_column=col_negoc, end_row=1, end_column=col_negoc + 2)
        c = ws.cell(row=1, column=col_negoc, value="NEGOCIAÇÃO")
        c.fill = _DARK; c.font = _WHITE; c.alignment = _CENTER; c.border = _BORDER

        # Linha 2 — nomes das lojas e sub-cabeçalhos de negociação
        cur = col_ini
        for loja in lojas:
            c = ws.cell(row=2, column=cur, value=_nome_col(loja))
            c.fill = _BROWN; c.font = _WHITE; c.alignment = _CENTER; c.border = _BORDER
            cur += 1

        for i, h in enumerate(["DESC %", "VALOR FECHADO", "BONIF %"]):
            c = ws.cell(row=2, column=col_negoc + i, value=h)
            c.fill = _DARK; c.font = _WHITE; c.alignment = _CENTER; c.border = _BORDER

        # Dados
        row_num = 3
        for item in itens:
            ws.cell(row=row_num, column=1, value=item['nome']).border = _BORDER
            ws.cell(row=row_num, column=2, value=item['caixa']).alignment = _CENTER

            cur = col_ini
            for loja in lojas:
                qtd = item['qtds'].get(loja.id, 0)
                c = ws.cell(row=row_num, column=cur, value=qtd)
                c.alignment = _CENTER; c.border = _BORDER
                cur += 1

            c = ws.cell(row=row_num, column=col_total, value=item['total'])
            c.font = _BOLD; c.alignment = _CENTER; c.border = _BORDER

            c = ws.cell(row=row_num, column=col_preco, value=item['preco'])
            c.number_format = '#,##0.00'; c.border = _BORDER

            desc = item.get('desconto', 0)
            c = ws.cell(row=row_num, column=col_negoc, value=desc)
            c.alignment = _CENTER; c.border = _BORDER; c.font = Font(color="0000FF", bold=True)

            preco_final = item['preco'] - (item['preco'] * (desc / 100))
            c = ws.cell(row=row_num, column=col_negoc + 1, value=preco_final)
            c.number_format = '#,##0.00'; c.font = Font(color="006400", bold=True); c.border = _BORDER

            c = ws.cell(row=row_num, column=col_negoc + 2, value=item.get('bonificacao', 0))
            c.alignment = _CENTER; c.border = _BORDER

            row_num += 1

        ws.column_dimensions['A'].width = 30
        for col in range(2, 20):
            try:
                ws.column_dimensions[chr(64 + col)].width = 12
            except Exception:
                pass

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _cell(ws, coord, value, fill):
    """Atalho para formatar célula de cabeçalho."""
    c = ws[coord]
    c.value = value
    c.fill = fill
    c.font = _WHITE
    c.alignment = _CENTER
    c.border = _BORDER
    return c

# ============================================================
# INSTRUÇÃO: Cole esta função ao FINAL do excel_service.py
# ============================================================

def gerar_excel_por_loja(periodo, fornecedores, dados):
    """
    Gera Excel: LOJAS nas linhas × FORNECEDORES nas colunas.
    Inclui coluna de Status e linha de totais com fórmulas.

    Parâmetros:
        periodo      — objeto Periodo
        fornecedores — lista de strings (nomes dos labs, ordenada)
        dados        — dict {nome_loja: {'forn': {forn: cx}, 'status': str}}

    Retorna BytesIO pronto para send_file().
    """
    from openpyxl.utils import get_column_letter

    _STATUS_FILL = {
        'Enviado':      PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid"),
        'Recebido':     PatternFill(start_color="0D47A1", end_color="0D47A1", fill_type="solid"),
        'Em andamento': PatternFill(start_color="E65100", end_color="E65100", fill_type="solid"),
        'Pendente':     PatternFill(start_color="B71C1C", end_color="B71C1C", fill_type="solid"),
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "Por Loja"

    n_forn         = len(fornecedores)
    status_col     = 2                      # coluna B = Status
    forn_start     = 3                      # fornecedores a partir de C
    total_col      = forn_start + n_forn    # última coluna = TOTAL
    data_start_row = 2
    data_end_row   = data_start_row + len(dados) - 1
    total_row      = data_end_row + 1

    # ── Cabeçalho ──────────────────────────────────────────────────────
    c = ws.cell(row=1, column=1, value="LOJA")
    c.font = _WHITE; c.fill = _BROWN; c.alignment = _CENTER; c.border = _BORDER

    c = ws.cell(row=1, column=status_col, value="STATUS")
    c.font = _WHITE; c.fill = _GRAY; c.alignment = _CENTER; c.border = _BORDER

    for ci, forn in enumerate(fornecedores, start=forn_start):
        c = ws.cell(row=1, column=ci, value=forn)
        c.font = _WHITE; c.fill = _BROWN
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _BORDER

    c = ws.cell(row=1, column=total_col, value="TOTAL")
    c.font = _WHITE; c.fill = _RED; c.alignment = _CENTER; c.border = _BORDER

    # ── Linhas de dados ─────────────────────────────────────────────────
    for ri, (loja_nome, info) in enumerate(dados.items(), start=data_start_row):
        # Loja
        c = ws.cell(row=ri, column=1, value=loja_nome)
        c.font = _BOLD; c.alignment = _CENTER; c.border = _BORDER

        # Status
        status = info.get('status', 'Pendente')
        c = ws.cell(row=ri, column=status_col, value=status)
        c.fill = _STATUS_FILL.get(status, _STATUS_FILL['Pendente'])
        c.font = _WHITE; c.alignment = _CENTER; c.border = _BORDER

        # Fornecedores
        forn_map = info.get('forn', {})
        for ci, forn in enumerate(fornecedores, start=forn_start):
            cx = forn_map.get(forn, 0)
            c  = ws.cell(row=ri, column=ci, value=cx)
            c.alignment = _CENTER; c.border = _BORDER
            if cx == 0:
                c.font = Font(color="CCCCCC")

        # Total da linha (fórmula)
        c1 = get_column_letter(forn_start)
        c2 = get_column_letter(forn_start + n_forn - 1)
        c  = ws.cell(row=ri, column=total_col,
                     value=f"=SUM({c1}{ri}:{c2}{ri})")
        c.font = Font(color="FFFFFF", bold=True)
        c.fill = _RED; c.alignment = _CENTER; c.border = _BORDER

    # ── Rodapé de totais ────────────────────────────────────────────────
    c = ws.cell(row=total_row, column=1, value="TOTAL")
    c.font = _WHITE; c.fill = _GRAY; c.alignment = _CENTER; c.border = _BORDER

    c = ws.cell(row=total_row, column=status_col, value="")
    c.fill = _GRAY; c.border = _BORDER

    for ci in range(forn_start, total_col + 1):
        col_letter = get_column_letter(ci)
        c = ws.cell(row=total_row, column=ci,
                    value=f"=SUM({col_letter}{data_start_row}:{col_letter}{data_end_row})")
        c.font = _WHITE; c.alignment = _CENTER; c.border = _BORDER
        c.fill = _RED if ci == total_col else _GRAY

    # ── Dimensões ────────────────────────────────────────────────────────
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions[get_column_letter(status_col)].width = 14
    ws.row_dimensions[1].height = 40
    for ci in range(forn_start, total_col + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 16

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output