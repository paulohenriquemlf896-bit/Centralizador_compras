"""
manage.py — CLI unificado do sistema Central de Compras
Substitui setup_db.py e msd.py com um único ponto de entrada.

Uso:
    python manage.py setup              # Cria banco, lojas, admin e grupos iniciais
    python manage.py importar           # Importa TODAS as abas configuradas
    python manage.py importar --aba MSD VALLÉE
    python manage.py importar --aba Agener Merck
    python manage.py importar --planilha outro_arquivo.xlsm
    python manage.py listar-abas        # Lista abas disponíveis na planilha
"""

import argparse
import os
import sys

from openpyxl import load_workbook
from werkzeug.security import generate_password_hash

# Importa o app e modelos
from app import app, db, Loja, Usuario, Grupo, Configuracao, Produto, Fornecedor

# ── Configurações padrão ──────────────────────────────────────────────────────

PLANILHA_PADRAO = 'criacao.xlsm'

# Abas importadas por padrão (quando --aba não é informado)
ABAS_PADRAO = ['Agener', 'Merck', 'Calbos', 'boehringer', 'Bravet', 'Elanco', 'MSD VALLÉE']

# Termos que indicam o fim da lista de produtos em qualquer aba
TERMOS_DE_PARADA = [
    "VALOR TOTAL", "OBSERVAÇ", "CONDICOES", "CONDIÇÕES",
    "DESC.", "PAGAMENTO", "PEDIDO MINIMO", "PEDIDO MÍNIMO", "TOTAL"
]


# ── Helpers ───────────────────────────────────────────────────────────────────

def _deve_parar(nome: str) -> bool:
    """Retorna True se a linha indica rodapé da planilha."""
    upper = nome.upper()
    return any(upper.startswith(t) for t in TERMOS_DE_PARADA)


def _ler_nome_produto(row) -> str | None:
    """Tenta ler o nome do produto da coluna A ou B."""
    val = None
    if row[0]:
        val = str(row[0]).strip()
    elif len(row) > 1 and row[1]:
        val = str(row[1]).strip()
    return val if val and val != 'None' else None


# ── Comandos ──────────────────────────────────────────────────────────────────

def cmd_setup():
    """Cria tabelas, dados iniciais de lojas, admin padrão e grupos."""
    with app.app_context():
        print("1. Criando tabelas...")
        db.create_all()

        print("2. Configurações iniciais...")
        if not Configuracao.query.filter_by(chave='admin_token').first():
            db.session.add(Configuracao(chave='admin_token', valor='123456'))

        if not Grupo.query.first():
            db.session.add_all([
                Grupo(nome="Medicamentos", codigo="2.1",      dia_limite=10),
                Grupo(nome="Perfumaria",   codigo="2.2",      dia_limite=17),
                Grupo(nome="Diversos",     codigo="DIVERSOS", dia_limite=1),
            ])
            print("   Grupos criados: Medicamentos, Perfumaria, Diversos")

        if not Loja.query.first():
            db.session.add_all([
                Loja(nome="Campina Grande"),
                Loja(nome="Maceió"),
                Loja(nome="Recife"),
                Loja(nome="Natal"),
                Loja(nome="Pesqueira"),
            ])
            print("   Lojas criadas: Campina Grande, Maceió, Recife, Natal, Pesqueira")

        if not Usuario.query.filter_by(email="admin@central.com").first():
            admin = Usuario(
                nome="Comprador",
                email="admin@central.com",
                senha=generate_password_hash("1234"),
                funcao="Admin",
                verificado=True,
            )
            db.session.add(admin)
            print("   Admin padrão criado: admin@central.com / senha: 1234")

        db.session.commit()
        print("\n✅ Setup concluído!")


def cmd_importar(abas: list[str], caminho_planilha: str):
    """Importa produtos de uma ou mais abas da planilha."""
    with app.app_context():
        if not os.path.exists(caminho_planilha):
            print(f"❌ Arquivo '{caminho_planilha}' não encontrado.")
            sys.exit(1)

        print(f"Abrindo: {caminho_planilha}")
        wb = load_workbook(caminho_planilha, data_only=True)

        # Normaliza nomes das abas disponíveis
        abas_disponiveis = {a.strip().lower(): a for a in wb.sheetnames}

        total_inseridos = 0
        total_abas = 0

        for nome_aba_alvo in abas:
            nome_normalizado = nome_aba_alvo.strip().lower()
            nome_real = abas_disponiveis.get(nome_normalizado)

            if not nome_real:
                print(f"  ⚠️  Aba '{nome_aba_alvo}' não encontrada na planilha. Pulando.")
                continue

            print(f"\n  → Processando aba: {nome_real}")
            sheet = wb[nome_real]
            fornecedor_nome = nome_real.strip()

            # Garante fornecedor no banco
            if not Fornecedor.query.filter_by(nome=fornecedor_nome).first():
                db.session.add(Fornecedor(nome=fornecedor_nome))
                db.session.commit()

            inseridos = 0
            for row in sheet.iter_rows(min_row=6, values_only=True):
                if not row:
                    continue

                nome_produto = _ler_nome_produto(row)
                if not nome_produto:
                    continue

                if _deve_parar(nome_produto):
                    print(f"     Fim detectado em: '{nome_produto}'")
                    break

                try:    caixa = int(row[2]) if len(row) > 2 and row[2] else 1
                except: caixa = 1

                try:    preco = float(row[3]) if len(row) > 3 and row[3] else 0.0
                except: preco = 0.0

                if not Produto.query.filter_by(nome=nome_produto, laboratorio=fornecedor_nome).first():
                    db.session.add(Produto(
                        nome=nome_produto,
                        laboratorio=fornecedor_nome,
                        grupo='2.1',
                        unidade_caixa=caixa,
                        preco=preco,
                    ))
                    inseridos += 1

            db.session.commit()
            print(f"     ✅ {inseridos} produto(s) inserido(s) de '{fornecedor_nome}'")
            total_inseridos += inseridos
            total_abas += 1

        print(f"\n✅ Importação concluída: {total_inseridos} produto(s) em {total_abas} aba(s).")


def cmd_listar_abas(caminho_planilha: str):
    """Lista as abas disponíveis na planilha."""
    if not os.path.exists(caminho_planilha):
        print(f"❌ Arquivo '{caminho_planilha}' não encontrado.")
        sys.exit(1)
    wb = load_workbook(caminho_planilha, data_only=True, read_only=True)
    print(f"Abas em '{caminho_planilha}':")
    for nome in wb.sheetnames:
        marcado = " ← padrão" if nome.strip().lower() in [a.lower() for a in ABAS_PADRAO] else ""
        print(f"  • {nome}{marcado}")


# ── Parser CLI ────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="CLI do sistema Central de Compras",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    sub = parser.add_subparsers(dest='comando', required=True)

    # setup
    sub.add_parser('setup', help='Cria banco, lojas, admin e grupos iniciais')

    # importar
    p_imp = sub.add_parser('importar', help='Importa produtos da planilha')
    p_imp.add_argument(
        '--aba', nargs='+', metavar='NOME',
        help='Nome(s) da(s) aba(s) a importar. Padrão: todas as abas configuradas.'
    )
    p_imp.add_argument(
        '--planilha', default=PLANILHA_PADRAO, metavar='ARQUIVO',
        help=f'Caminho da planilha .xlsm (padrão: {PLANILHA_PADRAO})'
    )

    # listar-abas
    p_lst = sub.add_parser('listar-abas', help='Lista abas disponíveis na planilha')
    p_lst.add_argument(
        '--planilha', default=PLANILHA_PADRAO, metavar='ARQUIVO',
        help=f'Caminho da planilha (padrão: {PLANILHA_PADRAO})'
    )

    args = parser.parse_args()

    if args.comando == 'setup':
        cmd_setup()

    elif args.comando == 'importar':
        abas = args.aba if args.aba else ABAS_PADRAO
        cmd_importar(abas, args.planilha)

    elif args.comando == 'listar-abas':
        cmd_listar_abas(args.planilha)


if __name__ == '__main__':
    main()
