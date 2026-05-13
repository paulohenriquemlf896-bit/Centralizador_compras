from flask import Flask, render_template, request, redirect, url_for, session, jsonify, send_file, abort
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime, date
import io
from flask_wtf.csrf import CSRFProtect
from services.excel_service import (
    gerar_excel_unico as _excel_unico,
    gerar_excel_consolidado,
    ordenar_lojas,
)
from services.email_service import enviar_email_pedido, montar_corpo_pedido
from openpyxl import load_workbook  # mantido para importar_produtos_excel
from werkzeug.security import generate_password_hash, check_password_hash
import random
import string
import base64 
import time
import os
from waitress import serve
import logging
from dotenv import load_dotenv
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address

def esconder_senha(texto):
    if not texto: return None
    return base64.b64encode(texto.encode("utf-8")).decode("utf-8")

def revelar_senha(texto_codificado):
    if not texto_codificado: return None
    try:
        return base64.b64decode(texto_codificado.encode("utf-8")).decode("utf-8")
    except:
        return texto_codificado

# Carrega as variáveis do arquivo .env
load_dotenv()

app = Flask(__name__)
limiter = Limiter(
    get_remote_address,
    app=app,
    default_limits=[],
    storage_uri="memory://"
)

@app.errorhandler(429)
def erro_muitas_tentativas(e):
    return redirect(url_for('login', erro='Muitas tentativas de login. Aguarde 1 minuto e tente novamente.'))
# Pode manter a chave ou buscar do .env também
app.secret_key = os.getenv('SECRET_KEY', 'chave_secreta_segura') 
csrf = CSRFProtect(app)

# --- CONFIGURAÇÃO DO BANCO DE DADOS ---
# Agora ele busca a URL do MariaDB no .env. Se não achar, usa o SQLite como fallback.
app.config['SQLALCHEMY_DATABASE_URI'] = os.getenv('DATABASE_URL', 'sqlite:///banco_dados.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

# --- MODELOS ---

class Configuracao(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    chave = db.Column(db.String(50), unique=True, nullable=False)
    valor = db.Column(db.String(200))

class Grupo(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(100), nullable=False)
    codigo = db.Column(db.String(50), unique=True, nullable=False)
    dia_limite = db.Column(db.Integer, nullable=False)

class Loja(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(100), nullable=False)

class Usuario(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(100), unique=True, nullable=False)
    senha = db.Column(db.String(200), nullable=False) 
    funcao = db.Column(db.String(20), nullable=False) 
    verificado = db.Column(db.Boolean, default=False)
    loja_id = db.Column(db.Integer, db.ForeignKey('loja.id'), nullable=True)
    loja = db.relationship('Loja', backref=db.backref('usuarios', lazy=True))

    smtp_server = db.Column(db.String(100), default='email-ssl.com.br')
    smtp_port = db.Column(db.Integer, default=465)
    smtp_email = db.Column(db.String(100))
    smtp_senha = db.Column(db.String(100))

class Periodo(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(100), nullable=False)      
    grupo_filtro = db.Column(db.String(50), nullable=False) 
    data_limite = db.Column(db.Date, nullable=False)
    mes = db.Column(db.Integer, nullable=False)
    ano = db.Column(db.Integer, nullable=False)
    ativo = db.Column(db.Boolean, default=True)

class Produto(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(200), nullable=False)
    laboratorio = db.Column(db.String(100), nullable=False)
    grupo = db.Column(db.String(50), default='2.1')       
    unidade_caixa = db.Column(db.Integer, default=1)
    preco = db.Column(db.Float, default=0.00) 

class Pedido(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    data_criacao = db.Column(db.DateTime, default=datetime.now)
    data_alteracao = db.Column(db.DateTime, default=datetime.now) 
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=False)
    loja_id = db.Column(db.Integer, db.ForeignKey('loja.id'), nullable=False)
    periodo_id = db.Column(db.Integer, db.ForeignKey('periodo.id'), nullable=False) 
    status = db.Column(db.String(20), default='Aberto') 

class ItemPedido(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    quantidade = db.Column(db.Integer, nullable=False)
    pedido_id = db.Column(db.Integer, db.ForeignKey('pedido.id'), nullable=False)
    produto_id = db.Column(db.Integer, db.ForeignKey('produto.id'), nullable=False)
    produto = db.relationship('Produto') 

class Negociacao(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    periodo_id = db.Column(db.Integer, db.ForeignKey('periodo.id'), nullable=False)
    produto_id = db.Column(db.Integer, db.ForeignKey('produto.id'), nullable=False)
    desconto = db.Column(db.Float, default=0.0)
    bonificacao = db.Column(db.Float, default=0.0)

class Observacao(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    periodo_id = db.Column(db.Integer, db.ForeignKey('periodo.id'), nullable=False)
    laboratorio = db.Column(db.String(100), nullable=False)
    loja_id = db.Column(db.Integer, db.ForeignKey('loja.id'), nullable=True) 
    texto = db.Column(db.Text, nullable=True)

class Fornecedor(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(100), unique=True, nullable=False)
    email = db.Column(db.String(100), nullable=True)

class LogAuditoria(db.Model):
    """Registra ações sensíveis: alteração de preço, desconto, envio de e-mail, etc."""
    id          = db.Column(db.Integer, primary_key=True)
    data_hora   = db.Column(db.DateTime, default=datetime.now, nullable=False)
    usuario_id  = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=True)
    usuario     = db.relationship('Usuario', backref=db.backref('logs', lazy=True))
    acao        = db.Column(db.String(80),  nullable=False)   # ex: 'PRECO_ALTERADO'
    entidade    = db.Column(db.String(80),  nullable=True)    # ex: 'Produto'
    entidade_id = db.Column(db.Integer,     nullable=True)    # ex: 42
    detalhe     = db.Column(db.Text,        nullable=True)    # JSON ou texto livre

class ProgressoLaboratorio(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    periodo_id = db.Column(db.Integer, db.ForeignKey('periodo.id'), nullable=False)
    loja_id = db.Column(db.Integer, db.ForeignKey('loja.id'), nullable=False)
    laboratorio = db.Column(db.String(100), nullable=False)
    concluido = db.Column(db.Boolean, default=False)

def registrar_log(acao, entidade=None, entidade_id=None, detalhe=None):
    """Helper: grava um registro de auditoria. Silencioso se não houver sessão ativa."""
    try:
        uid = session.get('usuario_id')
        log = LogAuditoria(
            usuario_id  = uid,
            acao        = acao,
            entidade    = entidade,
            entidade_id = int(entidade_id) if entidade_id else None,
            detalhe     = detalhe
        )
        db.session.add(log)
        # Não fazemos commit aqui — o commit já ocorre na rota chamadora
    except Exception as e:
        app.logger.warning(f'Falha ao registrar log de auditoria: {e}')

ultima_verificacao = None

@app.before_request
def verificar_datas_e_periodos():
    """
    FIX 10: Rotina diária robusta de manutenção de períodos.
    - Cada etapa tem seu próprio try/except com rollback isolado.
    - O admin é notificado no log de erro com mensagem descritiva.
    - ultima_verificacao só avança se AMBAS as etapas concluírem sem erro crítico.
    - Nunca deixa a sessão do banco em estado inconsistente.
    """
    global ultima_verificacao

    if request.endpoint and 'static' in request.endpoint:
        return

    hoje = date.today()
    if ultima_verificacao == hoje:
        return

    etapa1_ok = False
    etapa2_ok = False

    # ── Etapa 1: Fecha períodos vencidos ────────────────────────────────────
    try:
        periodos_abertos = Periodo.query.filter_by(ativo=True).all()
        fechados = 0
        for p in periodos_abertos:
            if hoje > p.data_limite:
                p.ativo = False
                fechados += 1
        if fechados:
            db.session.commit()
            app.logger.info(f"[Rotina] {fechados} período(s) fechado(s) automaticamente.")
        etapa1_ok = True
    except Exception as e:
        db.session.rollback()
        app.logger.error(
            f"[Rotina] FALHA ao fechar períodos vencidos: {e}. "
            "Verifique a conexão com o banco de dados."
        )

    # ── Etapa 2: Cria novos períodos baseado nos grupos ──────────────────────
    try:
        grupos_regras = Grupo.query.all()
        if not grupos_regras:
            etapa2_ok = True
        else:
            criados = 0
            for g in grupos_regras:
                dia_limite = g.dia_limite

                if hoje.day > dia_limite:
                    mes_alvo  = 1 if hoje.month == 12 else hoje.month + 1
                    ano_alvo  = hoje.year + 1 if hoje.month == 12 else hoje.year
                else:
                    mes_alvo, ano_alvo = hoje.month, hoje.year

                try:
                    data_alvo = date(ano_alvo, mes_alvo, dia_limite)
                except ValueError:
                    data_alvo = date(ano_alvo, mes_alvo, 28)

                nome_completo = f"{g.nome} {g.codigo} - {data_alvo.strftime('%b/%Y')}"

                if not Periodo.query.filter_by(
                    grupo_filtro=g.codigo, mes=mes_alvo, ano=ano_alvo
                ).first():
                    db.session.add(Periodo(
                        nome=nome_completo, grupo_filtro=g.codigo,
                        data_limite=data_alvo, mes=mes_alvo, ano=ano_alvo, ativo=True
                    ))
                    criados += 1

                    # Fecha o período anterior do mesmo grupo se ainda ativo
                    anterior = Periodo.query.filter(
                        Periodo.grupo_filtro == g.codigo,
                        Periodo.data_limite  <  data_alvo,
                        Periodo.ativo        == True
                    ).first()
                    if anterior:
                        anterior.ativo = False

            if criados:
                db.session.commit()
                app.logger.info(f"[Rotina] {criados} novo(s) período(s) criado(s).")
            etapa2_ok = True
    except Exception as e:
        db.session.rollback()
        app.logger.error(
            f"[Rotina] FALHA ao criar novos períodos: {e}. "
            "Verifique os grupos cadastrados em /admin/grupos."
        )

    # Só marca como verificado se ambas as etapas passaram sem erro crítico
    if etapa1_ok and etapa2_ok:
        ultima_verificacao = hoje
    else:
        app.logger.warning(
            "[Rotina] Verificação diária incompleta — será tentada novamente na próxima requisição."
        )

# --- ROTAS DE AUTENTICAÇÃO ---

@app.route('/')
def index():
    if 'usuario_id' in session:
        return redirect(url_for('dashboard_admin' if session.get('funcao') == 'Admin' else 'dashboard_loja'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
@limiter.limit("10 per minute", methods=["POST"], error_message="Muitas tentativas de login. Aguarde 1 minuto e tente novamente.")
def login():
    erro = request.args.get('erro')
    if request.method == 'POST':
        email = request.form['email'].lower().strip()
        senha = request.form['senha']
        user = Usuario.query.filter_by(email=email).first()
        
        senha_valida = False
        if user:
            if user.senha.startswith('scrypt:') or user.senha.startswith('pbkdf2:'):
                senha_valida = check_password_hash(user.senha, senha)
            else:
                senha_valida = (user.senha == senha)

            if senha_valida:
                if not user.verificado: 
                    erro = 'Sua conta ainda não foi aprovada pelo administrador.'
                else:
                    session['usuario_id'] = user.id
                    session['nome'] = user.nome
                    session['funcao'] = user.funcao
                    session['loja_id'] = user.loja_id 
                    return redirect(url_for('dashboard_admin' if user.funcao == 'Admin' else 'dashboard_loja'))
            else:
                erro = 'Email ou senha incorretos.'
        else:
            erro = 'Usuário não encontrado.'
            
    return render_template('login.html', erro=erro)

@app.route('/cadastro', methods=['GET', 'POST'])
def register():
    lojas = Loja.query.all()
    if request.method == 'POST':
        nome = request.form['nome']
        email = request.form['email'].lower().strip()
        senha = request.form['senha']
        token_input = request.form.get('token', '').strip()
        loja_id = request.form.get('loja_id')
        smtp_email = request.form.get('smtp_email')
        senha_plana = request.form.get('smtp_senha')
        smtp_senha = esconder_senha(senha_plana) if senha_plana else None

        if Usuario.query.filter_by(email=email).first(): 
            return render_template('register.html', lojas=lojas, erro="E-mail já existe!")
        
        config_token = Configuracao.query.filter_by(chave='admin_token').first()
        token_real = config_token.valor if config_token else 'TOKEN_INICIAL_123'
        
        funcao = 'Loja'
        if token_input:
            if token_input == token_real:
                funcao = 'Admin'
                loja_id = None
            else:
                return render_template('register.html', lojas=lojas, erro="Token de Administrador Inválido!")
        else:
            if not loja_id:
                return render_template('register.html', lojas=lojas, erro="Selecione sua loja ou insira um Token de Admin.")
            loja_id = int(loja_id)
            smtp_email = None
            smtp_senha = None

        senha_hash = generate_password_hash(senha)
            
        novo = Usuario(
            nome=nome, 
            email=email, 
            senha=senha_hash, 
            funcao=funcao, 
            loja_id=loja_id, 
            verificado=False,
            smtp_email=smtp_email, 
            smtp_senha=smtp_senha
        )
        db.session.add(novo)
        db.session.commit()
        return redirect(url_for('login', erro="Cadastro realizado! Aguarde aprovação."))
    
    return render_template('register.html', lojas=lojas)

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/dashboard/loja')
def dashboard_loja():
    if 'usuario_id' not in session: return redirect(url_for('login'))
    
    periodos_abertos = Periodo.query.filter_by(ativo=True).order_by(Periodo.data_limite).all()
    loja_id = session['loja_id']
    usuario_id = session['usuario_id']

    # Lógica Restaurada: Injeta status e cor dinamicamente
    for p in periodos_abertos:
        pedido = Pedido.query.filter_by(usuario_id=usuario_id, loja_id=loja_id, periodo_id=p.id).order_by(Pedido.id.desc()).first()
        
        p.status = "Pendente"
        p.cor = "danger"
        p.detalhe = "Toque para iniciar"
        p.pedido_id = None # Útil para link de recebimento
        
        if pedido:
            p.pedido_id = pedido.id
            if pedido.status == 'Recebido':
                p.status = "Recebido"
                p.cor = "secondary"
                p.detalhe = "Concluído"
            elif pedido.status == 'Enviado':
                p.status = "Enviado"
                p.cor = "success"
                p.detalhe = f"Enviado em {pedido.data_alteracao.strftime('%d/%m %H:%M')}"
            else:
                p.status = "Em Andamento"
                p.cor = "warning text-dark"
                p.detalhe = "Continuar editando..."
    
    return render_template('dashboard_loja.html', nome=session['nome'], periodos=periodos_abertos)

# --- ROTA QUE FALTAVA (CONFIRMAR CHEGADA) ---
@app.route('/pedido/confirmar_recebimento/<int:pedido_id>')
def confirmar_recebimento(pedido_id):
    if 'usuario_id' not in session: return redirect(url_for('login'))
    
    pedido = (db.session.get(Pedido, pedido_id) or abort(404))
    
    # Segurança: Só o dono do pedido pode confirmar
    if pedido.usuario_id != session['usuario_id']:
        return redirect(url_for('dashboard_loja'))
    
    if pedido.status == 'Enviado':
        pedido.status = 'Recebido'
        pedido.data_alteracao = datetime.now()
        db.session.commit()
    
    return redirect(url_for('dashboard_loja'))

@app.route('/loja/selecao_fabricante/<int:periodo_id>')
def selecao_fabricante(periodo_id):
    if 'usuario_id' not in session: return redirect(url_for('login'))
    
    periodo = (db.session.get(Periodo, periodo_id) or abort(404))
    loja_id = session['loja_id']
    
    # 1. BUSQUE O PEDIDO AQUI (UMA VEZ SÓ)
    # Isso garante que a variável 'pedido' sempre exista para o template
    pedido = Pedido.query.filter_by(loja_id=loja_id, periodo_id=periodo_id).first()
    
    labs_nomes = db.session.query(Produto.laboratorio).filter_by(grupo=periodo.grupo_filtro).distinct().order_by(Produto.laboratorio).all()
    
    lista_fabricantes = []
    for l in labs_nomes:
        nome_lab = l[0]
        status = 'nao-iniciado'
        
        # Verifica se o fabricante específico foi marcado como concluído
        progresso = ProgressoLaboratorio.query.filter_by(periodo_id=periodo_id, loja_id=loja_id, laboratorio=nome_lab).first()
        
        if progresso and progresso.concluido:
            status = 'concluido'
        elif pedido:
            # Se não está concluído, usa o 'pedido' que buscamos lá em cima para ver se há itens
            qtd = db.session.query(db.func.count(ItemPedido.id)).join(Produto).filter(
                ItemPedido.pedido_id == pedido.id, Produto.laboratorio == nome_lab
            ).scalar()
            if qtd > 0: 
                status = 'andamento'
                
        lista_fabricantes.append({'nome': nome_lab, 'status': status})
        
    # Agora 'pedido' está disponível com segurança aqui
    return render_template('fabricantes.html', 
                           periodo=periodo, 
                           lista_fabricantes=lista_fabricantes,
                           pedido=pedido)

@app.route('/loja/pedido/<int:periodo_id>/<path:laboratorio>', methods=['GET', 'POST'])
def pedido_form(periodo_id, laboratorio):
    if 'usuario_id' not in session: return redirect(url_for('login'))
    
    periodo = (db.session.get(Periodo, periodo_id) or abort(404))
    usuario_id = session['usuario_id']
    loja_id = session['loja_id']
    
    pedido = Pedido.query.filter_by(loja_id=loja_id, periodo_id=periodo_id).first()
    if not pedido:
        pedido = Pedido(usuario_id=usuario_id, loja_id=loja_id, periodo_id=periodo_id, status='Aberto')
        db.session.add(pedido)
        db.session.commit()
    
    bloqueado = (pedido.status != 'Aberto')

    if request.method == 'POST':
        if not bloqueado:
            # 1. Marca o laboratório como CONCLUÍDO para esta loja
            progresso = ProgressoLaboratorio.query.filter_by(
                periodo_id=periodo_id, loja_id=loja_id, laboratorio=laboratorio
            ).first()
            if not progresso:
                progresso = ProgressoLaboratorio(periodo_id=periodo_id, loja_id=loja_id, laboratorio=laboratorio)
                db.session.add(progresso)
            
            progresso.concluido = True
            pedido.data_alteracao = datetime.now()
            db.session.commit()
            
        return redirect(url_for('selecao_fabricante', periodo_id=periodo_id))

    # (O restante da busca de produtos continua igual)
    busca = request.args.get('busca', '').strip()
    query_prod = Produto.query.filter_by(grupo=periodo.grupo_filtro, laboratorio=laboratorio)
    if busca: query_prod = query_prod.filter(Produto.nome.ilike(f'%{busca}%'))
    produtos = query_prod.order_by(Produto.nome).all()
    itens_salvos = ItemPedido.query.filter_by(pedido_id=pedido.id).all()
    quantidades = {item.produto_id: item.quantidade for item in itens_salvos}
    total_produtos_lab = Produto.query.filter_by(grupo=periodo.grupo_filtro, laboratorio=laboratorio).count()

    return render_template('pedido_form.html', periodo=periodo, laboratorio=laboratorio, 
                           produtos=produtos, quantidades=quantidades, pedido=pedido,
                           bloqueado=bloqueado, busca=busca, total_produtos_lab=total_produtos_lab)

@app.route('/admin/importar_produtos_excel', methods=['POST'])
def importar_produtos_excel():
    if 'usuario_id' not in session or session.get('funcao') != 'Admin': 
        return redirect(url_for('login'))
    
    arquivo = request.files.get('arquivo_excel')
    grupo_padrao = request.form.get('grupo_padrao')

    if not arquivo or not grupo_padrao:
        return redirect(url_for('admin_produtos'))

    try:
        wb = load_workbook(arquivo, data_only=True)
        total_novos = 0
        total_atualizados = 0
        
        # --- LISTA DE PALAVRAS QUE INDICAM O FIM DA LISTA ---
        # Se a linha começar com qualquer uma dessas, o sistema PARA de ler a aba.
        termos_de_parada = [
            "VALOR TOTAL", 
            "OBSERVAÇ",   # Pega Observação e Observações
            "CONDICOES", 
            "CONDIÇÕES", 
            "DESC.", 
            "PAGAMENTO", 
            "PEDIDO MINIMO",
            "PEDIDO MÍNIMO",
            "TOTAL"
        ]

        for sheet in wb.worksheets:
            lab_nome = sheet.title.strip()
            
            # Cria Fornecedor
            forn = Fornecedor.query.filter_by(nome=lab_nome).first()
            if not forn:
                db.session.add(Fornecedor(nome=lab_nome))
                db.session.flush()

            print(f"Lendo aba: {lab_nome}...")

            for row in sheet.iter_rows(min_row=6, values_only=True):
                if not row: continue 
                
                nome_bruto = None
                
                # Tenta pegar Coluna A ou B
                if row[0]: nome_bruto = str(row[0]).strip()
                elif len(row) > 1 and row[1]: nome_bruto = str(row[1]).strip()
                
                if not nome_bruto: continue

                # === O FREIO DE MÃO (NOVO) ===
                # Converte para maiúsculo para comparar
                nome_upper = nome_bruto.upper()
                
                # Verifica se a linha começa com algum termo de parada
                deve_parar = False
                for termo in termos_de_parada:
                    if nome_upper.startswith(termo):
                        deve_parar = True
                        break
                
                if deve_parar:
                    print(f"    -> Parada acionada em: '{nome_bruto}'. Fim desta aba.")
                    break # SAI DO LOOP DESSA ABA (vai para a próxima aba se houver)
                # ==============================

                # Se passou pelo freio, é produto válido!
                produto = Produto.query.filter_by(nome=nome_bruto, laboratorio=lab_nome).first()
                
                if produto:
                    produto.grupo = grupo_padrao
                    total_atualizados += 1
                else:
                    novo = Produto(
                        nome=nome_bruto,
                        laboratorio=lab_nome,
                        grupo=grupo_padrao,
                        unidade_caixa=1,
                        preco=0.0
                    )
                    db.session.add(novo)
                    total_novos += 1

        db.session.commit()
        print(f"Sucesso! {total_novos} novos, {total_atualizados} atualizados.")
        
    except Exception as e:
        db.session.rollback()
        print(f"Erro na importação: {e}")
        
    return redirect(url_for('admin_produtos'))

@csrf.exempt
@app.route('/api/salvar_item_pedido', methods=['POST'])
def api_salvar_item_pedido():
    if 'usuario_id' not in session: return jsonify({'erro': 'Não logado'}), 401
    
    data = request.json
    periodo_id = data.get('periodo_id')
    produto_id = data.get('produto_id')
    quantidade = data.get('quantidade')
    
    try: quantidade = int(quantidade)
    except: quantidade = 0
    
    if not all([periodo_id, produto_id]): return jsonify({'erro': 'Dados incompletos'}), 400
    
    try:
        loja_id = session['loja_id']
        pedido = Pedido.query.filter_by(loja_id=loja_id, periodo_id=periodo_id).first()
        
        if not pedido:
            pedido = Pedido(usuario_id=session['usuario_id'], loja_id=loja_id, periodo_id=periodo_id, status='Aberto')
            db.session.add(pedido)
            db.session.flush() # Para ter o ID
            
        # === SEGURANÇA 2: TRAVA O AUTO-SAVE NA API ===
        if pedido.status != 'Aberto':
            return jsonify({'erro': 'Pedido Finalizado. Alterações bloqueadas.'}), 403

        # Salva item
        item = ItemPedido.query.filter_by(pedido_id=pedido.id, produto_id=produto_id).first()
        if item:
            if quantidade > 0: item.quantidade = quantidade
            else: db.session.delete(item)
        elif quantidade > 0:
            db.session.add(ItemPedido(pedido_id=pedido.id, produto_id=produto_id, quantidade=quantidade))
            
        pedido.data_alteracao = datetime.now()
        db.session.commit()
        return jsonify({'status': 'sucesso'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'erro': str(e)}), 500

@app.route('/loja/finalizar_pedido/<int:periodo_id>', methods=['POST'])
def finalizar_pedido_periodo(periodo_id):
    if 'usuario_id' not in session: return redirect(url_for('login'))
    loja_id = session['loja_id']
    pedido = Pedido.query.filter_by(loja_id=loja_id, periodo_id=periodo_id).first()
    if pedido:
        pedido.status = 'Enviado'
        pedido.data_alteracao = datetime.now()
        db.session.commit()
    return redirect(url_for('dashboard_loja'))

# ==============================================================================
# --- ROTAS ADMIN (GERENCIAMENTO) ---
# ==============================================================================

@app.route('/admin/produto/historico/<int:prod_id>')
def historico_preco(prod_id):
    if 'usuario_id' not in session or session.get('funcao') != 'Admin': 
        return redirect(url_for('login'))
    
    produto = db.session.get(Produto, prod_id)
    if not produto:
        return redirect(url_for('admin_produtos'))
        
    # Busca o histórico de alterações deste produto no Log de Auditoria
    logs = LogAuditoria.query.filter(
        LogAuditoria.entidade == 'Produto',
        LogAuditoria.entidade_id == prod_id
    ).order_by(LogAuditoria.data_hora.desc()).all()
    
    return render_template('historico_preco.html', produto=produto, logs=logs)

# ================================================================
# INSTRUÇÃO PRECISA:
#
# No app.py, SUBSTITUA APENAS este trecho (linhas 683 a 692):
#
#   @app.route('/admin/comparativo')
#   def comparativo_periodos():
#       ...
#       return render_template('comparativo_periodos.html', periodos=todos_periodos, lojas=lojas)
#
# PELO CÓDIGO ABAIXO.
# NÃO apague a rota pedidos_por_loja que está logo depois.
# ================================================================

@app.route('/admin/comparativo')
def comparativo_periodos():
    if 'usuario_id' not in session or session.get('funcao') != 'Admin':
        return redirect(url_for('login'))

    todos_periodos = Periodo.query.order_by(
        Periodo.ano.desc(), Periodo.mes.desc(), Periodo.id.desc()
    ).all()

    id_a = request.args.get('periodo_a', type=int)
    id_b = request.args.get('periodo_b', type=int)

    periodo_a   = None
    periodo_b   = None
    comparativo = []

    # Totais e contagens pré-calculados no Python
    # (evita comparar None com int no Jinja — causa do TypeError)
    resumo = {
        'total_qa': 0, 'total_qb': 0,
        'total_va': 0.0, 'total_vb': 0.0,
        'aumentaram': 0, 'reduziram': 0,
        'iguais': 0, 'novos': 0,
        'preco_dif': 0, 'zerados_b': 0,
        'var_qtd_pct': 0.0, 'var_val_pct': 0.0,
        'diff_val_total': 0.0,
    }

    if id_a and id_b and id_a != id_b:
        periodo_a = db.session.get(Periodo, id_a)
        periodo_b = db.session.get(Periodo, id_b)

    if periodo_a and periodo_b:

        def qtds_periodo(per):
            rows = db.session.query(
                Produto.id,
                db.func.sum(ItemPedido.quantidade).label('total')
            ).join(ItemPedido, ItemPedido.produto_id == Produto.id)\
             .join(Pedido, Pedido.id == ItemPedido.pedido_id)\
             .filter(Pedido.periodo_id == per.id)\
             .group_by(Produto.id).all()
            return {r[0]: int(r[1] or 0) for r in rows}

        def negs_periodo(per):
            return {n.produto_id: n for n in Negociacao.query.filter_by(periodo_id=per.id).all()}

        qtd_a = qtds_periodo(periodo_a)
        qtd_b = qtds_periodo(periodo_b)
        neg_a = negs_periodo(periodo_a)
        neg_b = negs_periodo(periodo_b)

        todos_ids = set(qtd_a.keys()) | set(qtd_b.keys())
        produtos  = {p.id: p for p in Produto.query.filter(Produto.id.in_(todos_ids)).all()}

        for pid, p in sorted(produtos.items(), key=lambda x: x[1].nome):
            qa = qtd_a.get(pid, 0)
            qb = qtd_b.get(pid, 0)

            # variacao é float ou None (novo); variacao_tipo é sempre string
            if qa > 0:
                variacao      = round(((qb - qa) / qa) * 100, 1)
                variacao_tipo = 'up' if variacao > 0 else ('down' if variacao < 0 else 'zero')
            elif qb > 0:
                variacao      = None
                variacao_tipo = 'novo'
            else:
                variacao      = 0.0
                variacao_tipo = 'zero'

            neg_a_item = neg_a.get(pid)
            neg_b_item = neg_b.get(pid)
            desc_a  = neg_a_item.desconto    if neg_a_item else 0.0
            desc_b  = neg_b_item.desconto    if neg_b_item else 0.0
            bonif_a = neg_a_item.bonificacao if neg_a_item else 0.0
            bonif_b = neg_b_item.bonificacao if neg_b_item else 0.0

            preco_a_val = p.preco
            preco_b_val = p.preco
            preco_mudou = round(preco_a_val, 2) != round(preco_b_val, 2)
            preco_liq_a = preco_a_val * (1 - desc_a / 100)
            preco_liq_b = preco_b_val * (1 - desc_b / 100)
            valor_a     = round(preco_liq_a * qa, 2)
            valor_b     = round(preco_liq_b * qb, 2)
            diff_val    = round(valor_b - valor_a, 2)

            comparativo.append({
                'id':            pid,
                'nome':          p.nome,
                'laboratorio':   p.laboratorio,
                'unidade_caixa': p.unidade_caixa,
                'qtd_a':         qa,
                'qtd_b':         qb,
                'variacao':      variacao,       # float ou None
                'variacao_tipo': variacao_tipo,  # 'up'|'down'|'zero'|'novo' — nunca None
                'preco_a':       preco_a_val,
                'preco_b':       preco_b_val,
                'preco_mudou':   preco_mudou,
                'desc_a':        desc_a,
                'desc_b':        desc_b,
                'bonif_a':       bonif_a,
                'bonif_b':       bonif_b,
                'preco_liq_a':   preco_liq_a,
                'preco_liq_b':   preco_liq_b,
                'valor_a':       valor_a,
                'valor_b':       valor_b,
                'diff_val':      diff_val,
            })

            # Acumula totais no Python (sem tocar no Jinja)
            resumo['total_qa'] += qa
            resumo['total_qb'] += qb
            resumo['total_va'] += valor_a
            resumo['total_vb'] += valor_b
            if variacao_tipo == 'up':    resumo['aumentaram'] += 1
            if variacao_tipo == 'down':  resumo['reduziram']  += 1
            if variacao_tipo == 'zero':  resumo['iguais']     += 1
            if variacao_tipo == 'novo':  resumo['novos']      += 1
            if preco_mudou:              resumo['preco_dif']  += 1
            if qb == 0:                  resumo['zerados_b']  += 1

        qa_t = resumo['total_qa']
        va_t = resumo['total_va']
        resumo['var_qtd_pct']    = round(((resumo['total_qb'] - qa_t) / qa_t * 100) if qa_t else 0, 1)
        resumo['var_val_pct']    = round(((resumo['total_vb'] - va_t) / va_t * 100) if va_t else 0, 1)
        resumo['diff_val_total'] = round(resumo['total_vb'] - resumo['total_va'], 2)

    return render_template(
        'comparativo_periodos.html',
        periodos=todos_periodos,
        periodo_a=periodo_a,
        periodo_b=periodo_b,
        id_a=id_a,
        id_b=id_b,
        comparativo=comparativo,
        resumo=resumo,
    )
@app.route('/admin/pedidos_por_loja')
def pedidos_por_loja():
    """
    Tabela resumo: LOJAS nas linhas × FORNECEDORES nas colunas.
    Melhorias incluídas:
      1. Destaque de lojas sem pedido de um fornecedor
      2. Status do pedido de cada loja (Enviado / Em Andamento / Pendente)
      3. Ordenação por coluna via query string ?ordenar=NomeFornecedor
      4. Filtro por fornecedores via checkboxes (?forn=X&forn=Y)
      5. Aviso de período aberto com lojas que ainda não enviaram
      6. Comparação com período anterior (▲▼)
    """
    if 'usuario_id' not in session or session.get('funcao') != 'Admin':
        return redirect(url_for('login'))

    periodo_id      = request.args.get('periodo_id', type=int)
    ordenar_por     = request.args.get('ordenar', '')          # nome do fornecedor ou 'total'
    ordenar_dir     = request.args.get('dir', 'desc')          # 'asc' | 'desc'
    forn_filtrados  = request.args.getlist('forn')             # lista de fornecedores selecionados

    ORDEM_LOJAS = ["Pesqueira", "Recife", "Campina Grande", "Natal", "Maceió"]
    todas_lojas = Loja.query.all()
    lojas = sorted(
        todas_lojas,
        key=lambda l: ORDEM_LOJAS.index(l.nome) if l.nome in ORDEM_LOJAS else 999
    )

    todos_periodos = Periodo.query.order_by(
        Periodo.ano.desc(), Periodo.mes.desc(), Periodo.id.desc()
    ).all()

    periodo         = None
    fornecedores    = []   # todos os labs do período
    forn_exibidos   = []   # labs após filtro de checkboxes
    tabela          = []
    totais_forn     = {}
    grand_total     = 0
    alertas         = []   # melhoria 5: lojas pendentes
    periodo_ant     = None # melhoria 6: período anterior
    totais_ant      = {}   # {loja_id: total_ant}

    if periodo_id:
        periodo = db.session.get(Periodo, periodo_id)

    if periodo:
        # ── Fornecedores do período ──────────────────────────────────────
        labs_q = db.session.query(Produto.laboratorio)\
            .filter_by(grupo=periodo.grupo_filtro)\
            .distinct().order_by(Produto.laboratorio).all()
        fornecedores = [l[0] for l in labs_q]

        # Aplica filtro de checkboxes (melhoria 4)
        forn_exibidos = [f for f in fornecedores if f in forn_filtrados] \
                        if forn_filtrados else fornecedores[:]

        # ── Quantidades: uma query só ────────────────────────────────────
        rows = db.session.query(
            Pedido.loja_id,
            Produto.laboratorio,
            db.func.sum(ItemPedido.quantidade).label('total')
        ).join(ItemPedido, ItemPedido.pedido_id == Pedido.id)\
         .join(Produto,    Produto.id == ItemPedido.produto_id)\
         .filter(
             Pedido.periodo_id == periodo.id,
             Produto.grupo     == periodo.grupo_filtro
         ).group_by(Pedido.loja_id, Produto.laboratorio).all()

        qtd_map = {(r.loja_id, r.laboratorio): int(r.total or 0) for r in rows}

        # ── Status dos pedidos por loja (melhoria 2) ─────────────────────
        pedidos_loja = {
            p.loja_id: p
            for p in Pedido.query.filter_by(periodo_id=periodo.id).all()
        }

        # ── Alertas de lojas sem envio (melhoria 5) ──────────────────────
        if periodo.ativo:
            for loja in lojas:
                ped = pedidos_loja.get(loja.id)
                if not ped or ped.status == 'Aberto':
                    nome_exib = 'Cruza' if loja.nome == 'Recife' else loja.nome
                    status_txt = 'Em andamento' if (ped and ped.status == 'Aberto') else 'Pendente'
                    alertas.append({'loja': nome_exib, 'status': status_txt})

        # ── Período anterior do mesmo grupo (melhoria 6) ─────────────────
        periodo_ant = Periodo.query.filter(
            Periodo.grupo_filtro == periodo.grupo_filtro,
            Periodo.id < periodo.id
        ).order_by(Periodo.id.desc()).first()

        if periodo_ant:
            rows_ant = db.session.query(
                Pedido.loja_id,
                db.func.sum(ItemPedido.quantidade).label('total')
            ).join(ItemPedido, ItemPedido.pedido_id == Pedido.id)\
             .join(Produto,    Produto.id == ItemPedido.produto_id)\
             .filter(
                 Pedido.periodo_id == periodo_ant.id,
                 Produto.grupo     == periodo_ant.grupo_filtro
             ).group_by(Pedido.loja_id).all()
            totais_ant = {r.loja_id: int(r.total or 0) for r in rows_ant}

        # ── Monta tabela ─────────────────────────────────────────────────
        totais_forn = {forn: 0 for forn in forn_exibidos}

        for loja in lojas:
            ped          = pedidos_loja.get(loja.id)
            total_ant_lj = totais_ant.get(loja.id, None)

            # Status (melhoria 2)
            if not ped:
                status_cor, status_icon, status_txt = 'danger',  '⏳', 'Pendente'
            elif ped.status == 'Enviado':
                status_cor, status_icon, status_txt = 'success', '✅', 'Enviado'
            elif ped.status == 'Recebido':
                status_cor, status_icon, status_txt = 'primary', '📦', 'Recebido'
            else:
                status_cor, status_icon, status_txt = 'warning', '✏️', 'Em andamento'

            linha = {
                'loja':           loja,
                'caixas_por_forn': {},
                'total_caixas':   0,
                'status_cor':     status_cor,
                'status_icon':    status_icon,
                'status_txt':     status_txt,
                'total_ant':      total_ant_lj,   # melhoria 6
            }

            for forn in forn_exibidos:
                cx = qtd_map.get((loja.id, forn), 0)
                linha['caixas_por_forn'][forn] = cx
                linha['total_caixas']         += cx
                totais_forn[forn]             += cx
                grand_total                   += cx

            tabela.append(linha)

        # ── Ordenação por coluna (melhoria 3) ────────────────────────────
        reverso = (ordenar_dir == 'desc')
        if ordenar_por == 'total':
            tabela.sort(key=lambda l: l['total_caixas'], reverse=reverso)
        elif ordenar_por == 'loja':
            tabela.sort(key=lambda l: l['loja'].nome, reverse=reverso)
        elif ordenar_por in forn_exibidos:
            tabela.sort(
                key=lambda l: l['caixas_por_forn'].get(ordenar_por, 0),
                reverse=reverso
            )

    return render_template(
        'pedidos_por_loja.html',
        todos_periodos=todos_periodos,
        periodo=periodo,
        periodo_id=periodo_id,
        fornecedores=fornecedores,
        forn_exibidos=forn_exibidos,
        forn_filtrados=forn_filtrados,
        lojas=lojas,
        tabela=tabela,
        totais_forn=totais_forn,
        grand_total=grand_total,
        alertas=alertas,
        periodo_ant=periodo_ant,
        ordenar_por=ordenar_por,
        ordenar_dir=ordenar_dir,
    )


@app.route('/admin/pedidos_por_loja/exportar/<int:periodo_id>')
def exportar_pedidos_por_loja(periodo_id):
    """Exporta a tabela de pedidos por loja como Excel."""
    if 'usuario_id' not in session or session.get('funcao') != 'Admin':
        return redirect(url_for('login'))

    periodo = (db.session.get(Periodo, periodo_id) or abort(404))

    ORDEM_LOJAS = ["Pesqueira", "Recife", "Campina Grande", "Natal", "Maceió"]
    todas_lojas = Loja.query.all()
    lojas = sorted(
        todas_lojas,
        key=lambda l: ORDEM_LOJAS.index(l.nome) if l.nome in ORDEM_LOJAS else 999
    )

    labs_q = db.session.query(Produto.laboratorio)\
        .filter_by(grupo=periodo.grupo_filtro)\
        .distinct().order_by(Produto.laboratorio).all()
    fornecedores = [l[0] for l in labs_q]

    rows = db.session.query(
        Pedido.loja_id,
        Produto.laboratorio,
        db.func.sum(ItemPedido.quantidade).label('total')
    ).join(ItemPedido, ItemPedido.pedido_id == Pedido.id)\
     .join(Produto,    Produto.id == ItemPedido.produto_id)\
     .filter(
         Pedido.periodo_id == periodo.id,
         Produto.grupo     == periodo.grupo_filtro
     ).group_by(Pedido.loja_id, Produto.laboratorio).all()

    qtd_map = {(r.loja_id, r.laboratorio): int(r.total or 0) for r in rows}

    pedidos_loja = {
        p.loja_id: p.status
        for p in Pedido.query.filter_by(periodo_id=periodo.id).all()
    }

    dados = {}
    for loja in lojas:
        nome_exibido = "Cruza" if loja.nome == "Recife" else loja.nome
        dados[nome_exibido] = {
            'forn':   {forn: qtd_map.get((loja.id, forn), 0) for forn in fornecedores},
            'status': pedidos_loja.get(loja.id, 'Pendente'),
        }

    from services.excel_service import gerar_excel_por_loja
    output       = gerar_excel_por_loja(periodo, fornecedores, dados)
    nome_arquivo = f"PedidosPorLoja_{periodo.nome.replace('/', '-')}.xlsx"

    return send_file(
        output,
        download_name=nome_arquivo,
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/admin/usuarios')
def admin_usuarios():
    if 'usuario_id' not in session or session.get('funcao') != 'Admin': return redirect(url_for('login'))
    
    config_token = Configuracao.query.filter_by(chave='admin_token').first()
    if not config_token:
        token_inicial = ''.join(random.choices(string.ascii_uppercase + string.digits, k=6))
        config_token = Configuracao(chave='admin_token', valor=token_inicial)
        db.session.add(config_token)
        db.session.commit()
    
    usuarios_pendentes = Usuario.query.filter_by(verificado=False).order_by(Usuario.id.desc()).all()
    usuarios_ativos = Usuario.query.filter_by(verificado=True).order_by(Usuario.nome).all()
    
    # FIX 8: Token NÃO é mais enviado direto ao template (visível no HTML fonte).
    # O JS agora busca via API autenticada /admin/api/rotacionar_token ao carregar.
    return render_template('admin_usuarios.html',
                           pendentes=usuarios_pendentes,
                           ativos=usuarios_ativos)

@app.route('/admin/gerar_novo_token')
def gerar_novo_token():
    if 'usuario_id' not in session or session.get('funcao') != 'Admin': return redirect(url_for('login'))
    novo_token = ''.join(random.choices(string.ascii_uppercase + string.digits, k=8))
    conf = Configuracao.query.filter_by(chave='admin_token').first()
    if conf: conf.valor = novo_token
    else: db.session.add(Configuracao(chave='admin_token', valor=novo_token))
    db.session.commit()
    return redirect(url_for('admin_usuarios'))

# --- ROTA API PARA O JAVASCRIPT ---
@csrf.exempt
@app.route('/admin/api/rotacionar_token')
def api_rotacionar_token():
    if 'usuario_id' not in session or session.get('funcao') != 'Admin': 
        return jsonify({'erro': 'Não autorizado'}), 401
    
    # Gera um novo token de 6 caracteres (Letras + Números)
    novo_token = ''.join(random.choices(string.ascii_uppercase + string.digits, k=6))
    
    # Atualiza no Banco
    conf = Configuracao.query.filter_by(chave='admin_token').first()
    if conf:
        conf.valor = novo_token
    else:
        db.session.add(Configuracao(chave='admin_token', valor=novo_token))
    
    db.session.commit()
    
    return jsonify({'token': novo_token})

@app.route('/admin/aprovar/<int:user_id>')
def aprovar_usuario(user_id):
    if 'usuario_id' not in session or session.get('funcao') != 'Admin': return redirect(url_for('login'))
    user = db.session.get(Usuario, user_id)
    if user:
        user.verificado = True
        db.session.commit()
    return redirect(url_for('admin_usuarios'))

@app.route('/admin/excluir/<int:user_id>')
def excluir_usuario(user_id):
    if 'usuario_id' not in session or session.get('funcao') != 'Admin': return redirect(url_for('login'))
    if user_id == session['usuario_id']: return redirect(url_for('admin_usuarios')) 
    user = db.session.get(Usuario, user_id)
    if user:
        db.session.delete(user)
        db.session.commit()
    return redirect(url_for('admin_usuarios'))

# --- ROTAS DE GRUPOS ---

# --- ROTAS DE GRUPOS ---

@app.route('/admin/grupos', methods=['GET', 'POST'])
def admin_grupos():
    # 1. ADICIONE ESTA LINHA PARA PUXAR A VARIÁVEL GLOBAL
    global ultima_verificacao 

    if 'usuario_id' not in session or session.get('funcao') != 'Admin': return redirect(url_for('login'))
    
    if request.method == 'POST':
        nome = request.form.get('nome')
        codigo = request.form.get('codigo')
        try: dia = int(request.form.get('dia_limite'))
        except: dia = 10
        
        if nome and codigo:
            existe = Grupo.query.filter_by(codigo=codigo).first()
            if not existe:
                db.session.add(Grupo(nome=nome, codigo=codigo, dia_limite=dia))
                db.session.commit()
                
                # 2. ADICIONE ESTA LINHA AQUI!
                # Isso força o sistema a recriar os períodos no exato segundo em que a página recarregar.
                ultima_verificacao = None 
                
    grupos = Grupo.query.order_by(Grupo.dia_limite).all()
    return render_template('admin_grupos.html', grupos=grupos)

@app.route('/admin/grupo/editar/<int:grupo_id>', methods=['POST'])
def editar_grupo(grupo_id):
    global ultima_verificacao

    if 'usuario_id' not in session or session.get('funcao') != 'Admin':
        return redirect(url_for('login'))

    g = db.session.get(Grupo, grupo_id)
    if g:
        novo_nome   = request.form.get('nome', '').strip()
        novo_codigo = request.form.get('codigo', '').strip()
        try:
            novo_dia = int(request.form.get('dia_limite'))
        except (TypeError, ValueError):
            novo_dia = g.dia_limite

        if novo_nome and novo_codigo:
            codigo_antigo = g.codigo

            # Se o código mudou, migra produtos e períodos vinculados
            if novo_codigo != codigo_antigo:
                conflito = Grupo.query.filter(
                    Grupo.codigo == novo_codigo, Grupo.id != grupo_id
                ).first()
                if conflito:
                    return redirect(url_for('admin_grupos'))

                Produto.query.filter_by(grupo=codigo_antigo).update({'grupo': novo_codigo})
                Periodo.query.filter_by(grupo_filtro=codigo_antigo).update({'grupo_filtro': novo_codigo})
                codigo_alvo = novo_codigo
            else:
                codigo_alvo = codigo_antigo

            # Se o dia mudou, recalcula data_limite dos períodos ativos deste grupo
            if novo_dia != g.dia_limite:
                periodos_ativos = Periodo.query.filter_by(grupo_filtro=codigo_alvo, ativo=True).all()
                for p in periodos_ativos:
                    try:
                        nova_data = date(p.ano, p.mes, novo_dia)
                    except ValueError:
                        nova_data = date(p.ano, p.mes, 28)
                    p.data_limite = nova_data

            g.nome       = novo_nome
            g.codigo     = novo_codigo
            g.dia_limite = novo_dia
            db.session.commit()
            ultima_verificacao = None

    return redirect(url_for('admin_grupos'))

@app.route('/admin/grupo/excluir/<int:grupo_id>')
def excluir_grupo(grupo_id):
    global ultima_verificacao 

    if 'usuario_id' not in session or session.get('funcao') != 'Admin': 
        return redirect(url_for('login'))
        
    g = db.session.get(Grupo, grupo_id)
    if g:
        codigo_grupo = g.codigo
        
        # 1. Identificar todos os produtos que pertencem a este grupo
        produtos_do_grupo = Produto.query.filter_by(grupo=codigo_grupo).all()
        ids_produtos = [p.id for p in produtos_do_grupo]

        if ids_produtos:
            # 2. Apagar referências em Itens de Pedido e Negociações para evitar erros de banco
            ItemPedido.query.filter(ItemPedido.produto_id.in_(ids_produtos)).delete(synchronize_session=False)
            Negociacao.query.filter(Negociacao.produto_id.in_(ids_produtos)).delete(synchronize_session=False)

            # 3. Apagar os produtos do grupo de uma vez
            Produto.query.filter_by(grupo=codigo_grupo).delete(synchronize_session=False)

        # 4. Desativar períodos associados para limparem o dashboard
        Periodo.query.filter_by(grupo_filtro=codigo_grupo).update({"ativo": False})
            
        # 5. Apagar o registro do grupo em si
        db.session.delete(g)
        
        db.session.commit()
        
        # Força o sistema a reprocessar os períodos ativos na próxima requisição
        ultima_verificacao = None 
        
    return redirect(url_for('admin_grupos'))

# --- ROTAS DE FORNECEDORES (E-mails e Nomes) ---

@app.route('/admin/fornecedores', methods=['GET', 'POST'])
def admin_fornecedores():
    if 'usuario_id' not in session or session.get('funcao') != 'Admin': return redirect(url_for('login'))
    
    # Adicionar Fornecedor
    if request.method == 'POST':
        nome = request.form.get('nome')
        if nome:
            if not Fornecedor.query.filter_by(nome=nome).first():
                db.session.add(Fornecedor(nome=nome))
                db.session.commit()
    
    fornecedores = Fornecedor.query.order_by(Fornecedor.nome).all()
    return render_template('admin_fornecedores.html', fornecedores=fornecedores)

@app.route('/admin/fornecedor/excluir/<int:id>')
def excluir_fornecedor(id):
    if 'usuario_id' not in session or session.get('funcao') != 'Admin': return redirect(url_for('login'))
    f = db.session.get(Fornecedor, id)
    if f:
        try:
            db.session.delete(f)
            db.session.commit()
        except:
            db.session.rollback() 
    return redirect(url_for('admin_fornecedores'))

@app.route('/admin/salvar_email_fornecedor', methods=['POST'])
def salvar_email_fornecedor():
    if 'usuario_id' not in session or session.get('funcao') != 'Admin': return redirect(url_for('login'))
    fornecedor_id = request.form.get('fornecedor_id')
    email = request.form.get('email')
    forn = db.session.get(Fornecedor, fornecedor_id)
    if forn:
        forn.email = email
        db.session.commit()
    return redirect(url_for('admin_fornecedores'))

# --- ROTAS DE PRODUTOS ---

@app.route('/admin/produtos', methods=['GET', 'POST'])
def admin_produtos():
    if 'usuario_id' not in session or session.get('funcao') != 'Admin': return redirect(url_for('login'))
    
    if request.method == 'POST':
        nome = request.form.get('nome')
        laboratorio = request.form.get('laboratorio') 
        grupo = request.form.get('grupo') 
        try: unidade_caixa = int(request.form.get('unidade_caixa'))
        except: unidade_caixa = 1
        try: preco = float(request.form.get('preco'))
        except: preco = 0.0
        
        if nome and laboratorio:
            existe = Produto.query.filter_by(nome=nome, laboratorio=laboratorio).first()
            if not existe:
                db.session.add(Produto(nome=nome, laboratorio=laboratorio, grupo=grupo, unidade_caixa=unidade_caixa, preco=preco))
                db.session.commit()

    # FIX 5: Paginação — carrega apenas 50 por vez em vez de tudo de uma vez
    pagina      = request.args.get('pagina', 1, type=int)
    por_pagina  = 50
    filtro_nome = request.args.get('q', '').strip()
    filtro_lab  = request.args.get('lab', '').strip()
    filtro_grp  = request.args.get('grp', '').strip()

    query = Produto.query
    if filtro_nome:
        query = query.filter(Produto.nome.ilike(f'%{filtro_nome}%'))
    if filtro_lab:
        query = query.filter(Produto.laboratorio == filtro_lab)
    if filtro_grp:
        query = query.filter(Produto.grupo == filtro_grp)

    paginacao = query.order_by(Produto.laboratorio, Produto.nome).paginate(
        page=pagina, per_page=por_pagina, error_out=False
    )

    laboratorios_cadastrados = Fornecedor.query.order_by(Fornecedor.nome).all()
    grupos_cadastrados        = Grupo.query.order_by(Grupo.nome).all()

    return render_template('admin_produtos.html',
                           produtos=paginacao.items,
                           paginacao=paginacao,
                           filtro_nome=filtro_nome,
                           filtro_lab=filtro_lab,
                           filtro_grp=filtro_grp,
                           laboratorios_cadastrados=laboratorios_cadastrados,
                           grupos_cadastrados=grupos_cadastrados)

@app.route('/admin/produto/editar', methods=['POST'])
def editar_produto():
    if 'usuario_id' not in session or session.get('funcao') != 'Admin': return redirect(url_for('login'))
    
    prod_id = request.form.get('prod_id')
    prod = db.session.get(Produto, prod_id)
    
    if prod:
        preco_anterior = prod.preco
        prod.nome       = request.form.get('nome')
        prod.laboratorio= request.form.get('laboratorio')
        prod.grupo      = request.form.get('grupo')
        try: prod.unidade_caixa = int(request.form.get('unidade_caixa'))
        except: pass
        try:
            novo_preco = float(request.form.get('preco'))
            prod.preco = novo_preco
        except: novo_preco = prod.preco

        # FIX 9: Log de auditoria para alteração de produto
        registrar_log(
            acao='PRODUTO_EDITADO',
            entidade='Produto', entidade_id=prod.id,
            detalhe=f'Preço: {preco_anterior:.2f} → {novo_preco:.2f} | Lab: {prod.laboratorio}'
        )
        db.session.commit()

    return redirect(url_for('admin_produtos'))

@app.route('/admin/produto/excluir/<int:prod_id>')
def excluir_produto(prod_id):
    if 'usuario_id' not in session or session.get('funcao') != 'Admin': return redirect(url_for('login'))
    prod = db.session.get(Produto, prod_id)
    if prod:
        try:
            db.session.delete(prod)
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            app.logger.error(f'Erro ao excluir produto {prod_id}: {e}')
    return redirect(url_for('admin_produtos'))

# --- ROTAS DE CONSOLIDAÇÃO E ENVIO (ADMIN) ---
@app.route('/dashboard/admin')
def dashboard_admin():
    if 'usuario_id' not in session: return redirect(url_for('login'))

    nomes = {1:'Jan',2:'Fev',3:'Mar',4:'Abr',5:'Mai',6:'Jun',
             7:'Jul',8:'Ago',9:'Set',10:'Out',11:'Nov',12:'Dez'}

    # ── Opções de meses para o filtro (1 query) ───────────────────────────
    todos_periodos = Periodo.query.order_by(Periodo.ano.desc(), Periodo.mes.desc()).all()
    opcoes_meses, meses_vistos = [], set()
    for p in todos_periodos:
        k = f"{p.mes}-{p.ano}"
        if k not in meses_vistos:
            opcoes_meses.append({'valor': k, 'texto': f"{nomes[p.mes]}/{p.ano}"})
            meses_vistos.add(k)

    # ── Períodos filtrados (1 query) ──────────────────────────────────────
    filtro = request.args.get('filtro_data')
    if filtro:
        mes_f, ano_f = filtro.split('-')
        base_q = Periodo.query.filter_by(mes=int(mes_f), ano=int(ano_f))
    else:
        base_q = Periodo.query.filter_by(ativo=True)
    periodos_filt = base_q.order_by(Periodo.ano.desc(), Periodo.mes.desc(), Periodo.data_limite).all()

    if not periodos_filt:
        return render_template('dashboard_admin.html', nome=session['nome'],
                               dados_painel=[], opcoes_meses=opcoes_meses,
                               filtro_atual=filtro, data_hoje=datetime.now().strftime('%d/%m/%Y'))

    ids_periodos = [p.id for p in periodos_filt]
    todas_lojas  = Loja.query.order_by(Loja.nome).all()

    # ── 1 query: todos os pedidos dos períodos visíveis ───────────────────
    todos_pedidos = Pedido.query.filter(Pedido.periodo_id.in_(ids_periodos)).all()
    # índice: (periodo_id, loja_id) -> pedido mais recente
    ped_idx = {}
    for ped in todos_pedidos:
        chave = (ped.periodo_id, ped.loja_id)
        if chave not in ped_idx or ped.id > ped_idx[chave].id:
            ped_idx[chave] = ped

    # ── 1 query: contagem de itens por pedido (só para pedidos Aberto) ────
    ids_abertos = [ped.id for ped in todos_pedidos if ped.status == 'Aberto']
    contagem_itens = {}
    if ids_abertos:
        rows = db.session.query(ItemPedido.pedido_id, db.func.count(ItemPedido.id))            .filter(ItemPedido.pedido_id.in_(ids_abertos))            .group_by(ItemPedido.pedido_id).all()
        contagem_itens = {r[0]: r[1] for r in rows}

    # ── 1 query: total de itens pedidos por (produto_id, periodo_id) ──────
    totais_lab = db.session.query(
        Produto.laboratorio,
        Produto.grupo,
        db.func.count(Produto.id).label('qtd_produtos'),
        db.func.coalesce(db.func.sum(ItemPedido.quantidade), 0).label('total_pedidos')
    ).outerjoin(ItemPedido, ItemPedido.produto_id == Produto.id)     .outerjoin(Pedido, db.and_(
         Pedido.id == ItemPedido.pedido_id,
         Pedido.periodo_id.in_(ids_periodos)
     ))     .group_by(Produto.laboratorio, Produto.grupo).all()

    # índice: grupo -> {lab_nome -> {produtos_count, total_pedidos}}
    labs_por_grupo = {}
    for row in totais_lab:
        g = row.grupo
        if g not in labs_por_grupo:
            labs_por_grupo[g] = {}
        labs_por_grupo[g][row.laboratorio] = {
            'nome': row.laboratorio,
            'produtos_count': row.qtd_produtos,
            'total_pedidos': int(row.total_pedidos),
        }

    # ── Monta dados_painel sem nenhuma query adicional ────────────────────
    dados_painel = []
    for p in periodos_filt:
        # Status de cada loja
        sl = []
        for loja in todas_lojas:
            ped = ped_idx.get((p.id, loja.id))
            inf = {'nome': loja.nome, 'cor': 'danger', 'texto': '⏳', 'detalhe': 'Pendente'}
            if ped:
                if ped.status == 'Recebido':
                    inf.update({'cor': 'primary', 'texto': '📦', 'detalhe': 'Recebido na Loja'})
                elif ped.status == 'Enviado':
                    inf.update({'cor': 'success', 'texto': '✅',
                                'detalhe': f"Enviado {ped.data_alteracao.strftime('%d/%m')}"})
                elif ped.status == 'Aberto':
                    if contagem_itens.get(ped.id, 0) > 0:
                        inf.update({'cor': 'warning text-dark', 'texto': '✏️', 'detalhe': 'Editando'})
            sl.append(inf)

        # Fabricantes do grupo
        fabricantes_data = sorted(
            labs_por_grupo.get(p.grupo_filtro, {}).values(),
            key=lambda x: x['nome']
        )

        dados_painel.append({'periodo': p, 'lojas': sl, 'fabricantes': fabricantes_data})

    return render_template('dashboard_admin.html', nome=session['nome'], dados_painel=dados_painel,
                           opcoes_meses=opcoes_meses, filtro_atual=filtro,
                           data_hoje=datetime.now().strftime('%d/%m/%Y'))

@app.route('/admin/consolidacao')
@app.route('/admin/consolidacao/<int:periodo_id>')
def consolidacao_pedidos(periodo_id=None):
    if 'usuario_id' not in session or session.get('funcao') != 'Admin':
        return redirect(url_for('login'))

    if periodo_id:
        periodo_ativo = (db.session.get(Periodo, periodo_id) or abort(404))
    else:
        periodo_ativo = Periodo.query.filter_by(ativo=True).order_by(Periodo.data_limite).first()
        if not periodo_ativo:
            periodo_ativo = Periodo.query.order_by(Periodo.id.desc()).first()
    if not periodo_ativo:
        return "Nenhum período encontrado."

    ordem_personalizada = ["Pesqueira", "Recife", "Campina Grande", "Natal", "Maceió"]
    todas_lojas = Loja.query.all()
    lojas = sorted(todas_lojas, key=lambda l: ordem_personalizada.index(l.nome) if l.nome in ordem_personalizada else 999)

    # Envia só a lista de laboratórios para montar as abas — SEM os dados pesados
    laboratorios = db.session.query(Produto.laboratorio)\
        .filter_by(grupo=periodo_ativo.grupo_filtro)\
        .distinct().order_by(Produto.laboratorio).all()
    laboratorios = [l[0] for l in laboratorios]

    return render_template('consolidacao.html',
                           laboratorios=laboratorios,
                           periodo=periodo_ativo,
                           lojas=lojas)



@app.route('/admin/consolidacao/aba/<int:periodo_id>/<path:laboratorio>')
def consolidacao_aba(periodo_id, laboratorio):
    """Retorna o HTML de UMA aba — chamado via fetch pelo frontend."""
    if 'usuario_id' not in session or session.get('funcao') != 'Admin':
        return '', 401

    periodo = (db.session.get(Periodo, periodo_id) or abort(404))
    
    # Ordem personalizada das lojas para a visualização na matriz
    ordem_personalizada = ["Pesqueira", "Recife", "Campina Grande", "Natal", "Maceió"]
    todas_lojas = Loja.query.all()
    lojas = sorted(todas_lojas, key=lambda l: ordem_personalizada.index(l.nome) if l.nome in ordem_personalizada else 999)

    # Busca todos os produtos do laboratório que pertencem ao grupo do período ativo
    produtos = Produto.query.filter_by(grupo=periodo.grupo_filtro, laboratorio=laboratorio)\
        .order_by(Produto.nome).all()

    # Busca negociações e quantidades de uma vez para evitar múltiplas consultas ao banco (N+1)
    ids_produtos = [p.id for p in produtos]

    negociacoes = {n.produto_id: n for n in
                   Negociacao.query.filter_by(periodo_id=periodo.id)
                   .filter(Negociacao.produto_id.in_(ids_produtos)).all()}

    qtd_map = {}
    rows = db.session.query(ItemPedido.produto_id, Pedido.loja_id,
                            db.func.sum(ItemPedido.quantidade))\
        .join(Pedido)\
        .filter(Pedido.periodo_id == periodo.id,
                ItemPedido.produto_id.in_(ids_produtos))\
        .group_by(ItemPedido.produto_id, Pedido.loja_id).all()
        
    for prod_id, loja_id, soma in rows:
        qtd_map[(prod_id, loja_id)] = soma or 0

    lista_itens = []
    for p in produtos:
        negoc = negociacoes.get(p.id)
        qtds_lojas = {loja.id: qtd_map.get((p.id, loja.id), 0) for loja in lojas}
        desconto = negoc.desconto if negoc else 0.0
        total_unid = sum(qtds_lojas.values())
        
        # FIX: Conversão para float evita o erro TypeError com tipos Decimal do banco
        # Além de calcular o valor total, agora calculamos o preço por unidade (líquido)
        preco_liquido = float(p.preco) * (1 - float(desconto) / 100)
        valor_fechado = preco_liquido * float(total_unid)
        
        lista_itens.append({
            'id': p.id, 
            'nome': p.nome, 
            'caixa': p.unidade_caixa, 
            'preco': p.preco,
            'qtds': qtds_lojas, 
            'total': total_unid,
            'desconto': desconto,
            'bonificacao': negoc.bonificacao if negoc else 0.0,
            'preco_liquido': preco_liquido,  # Informação do preço por unidade
            'valor_fechado': valor_fechado,  # Valor total da linha
        })

    # Resumos calculados no servidor para exibição no rodapé (tfoot)
    resumo_por_loja = {loja.id: sum(item['qtds'].get(loja.id, 0) for item in lista_itens) for loja in lojas}
    resumo_total_geral = sum(item['total'] for item in lista_itens)
    resumo_valor_fechado = sum(item['valor_fechado'] for item in lista_itens)

    # Busca observações vinculadas a este laboratório e período
    obs_db = Observacao.query.filter_by(periodo_id=periodo.id, laboratorio=laboratorio).all()
    obs_map = {f"{o.laboratorio}|{o.loja_id if o.loja_id else 'GERAL'}": o.texto for o in obs_db}

    return render_template('_aba_consolidacao.html',
                        laboratorio=laboratorio,
                        lista_itens=lista_itens,
                        lojas=lojas,
                        periodo=periodo,
                        obs_map=obs_map,
                        resumo_por_loja=resumo_por_loja,
                        resumo_total_geral=resumo_total_geral,
                        resumo_valor_fechado=resumo_valor_fechado)

@csrf.exempt
@app.route('/admin/salvar_consolidacao', methods=['POST'])
def salvar_consolidacao():
    data = request.json
    periodo_id = data.get('periodo_id')
    if not periodo_id: return jsonify({'erro': 'Erro ID'}), 400
    try:
        for item in data.get('produtos', []):
            prod_id = item['produto_id']
            try: desc = float(item['desconto'])
            except: desc = 0.0
            try: bonif = float(item['bonificacao'])
            except: bonif = 0.0
            try: novo_preco = float(item['preco'])
            except: novo_preco = 0.0
            prod = db.session.get(Produto, prod_id)
            if prod and novo_preco > 0:
                # FIX 7: Registra se o preço de tabela foi alterado
                if abs(prod.preco - novo_preco) > 0.001:
                    # FIX 9b: Log da alteração de preço na consolidação
                    registrar_log(
                        acao='PRECO_ALTERADO',
                        entidade='Produto', entidade_id=prod.id,
                        detalhe=f'Período {periodo_id} | {prod.nome}: R${prod.preco:.2f} → R${novo_preco:.2f}'
                    )
                prod.preco = novo_preco
            negoc = Negociacao.query.filter_by(periodo_id=periodo_id, produto_id=prod_id).first()
            if not negoc:
                negoc = Negociacao(periodo_id=periodo_id, produto_id=prod_id)
                db.session.add(negoc)
            # FIX 9b: Log de desconto/bonificação
            if negoc.desconto != desc or negoc.bonificacao != bonif:
                registrar_log(
                    acao='NEGOCIACAO_ALTERADA',
                    entidade='Negociacao', entidade_id=prod_id,
                    detalhe=f'Período {periodo_id} | Desc: {negoc.desconto}%→{desc}% | Bonif: {negoc.bonificacao}%→{bonif}%'
                )
            negoc.desconto = desc
            negoc.bonificacao = bonif
            for loja_id_str, qtd_val in item.get('qtds', {}).items():
                loja_id = int(loja_id_str)
                try: qtd_nova = int(qtd_val)
                except: qtd_nova = 0
                pedido = Pedido.query.filter_by(loja_id=loja_id, periodo_id=periodo_id).first()
                if not pedido and qtd_nova > 0:
                    dono = Usuario.query.filter_by(loja_id=loja_id).first()
                    uid = dono.id if dono else session['usuario_id']
                    pedido = Pedido(usuario_id=uid, loja_id=loja_id, periodo_id=periodo_id, status='Aberto')
                    db.session.add(pedido)
                    db.session.flush()
                if pedido:
                    item_ped = ItemPedido.query.filter_by(pedido_id=pedido.id, produto_id=prod_id).first()
                    if item_ped:
                        if qtd_nova > 0: item_ped.quantidade = qtd_nova
                        else: db.session.delete(item_ped)
                    elif qtd_nova > 0: db.session.add(ItemPedido(pedido_id=pedido.id, produto_id=prod_id, quantidade=qtd_nova))
        for obs in data.get('observacoes', []):
            lab = obs['laboratorio']
            loja_id_obs = int(obs['loja_id']) if obs['loja_id'] else None
            texto = obs['texto']
            filtro = Observacao.loja_id == loja_id_obs if loja_id_obs else Observacao.loja_id.is_(None)
            reg = Observacao.query.filter(Observacao.periodo_id == periodo_id, Observacao.laboratorio == lab, filtro).first()
            if reg: reg.texto = texto
            elif texto: db.session.add(Observacao(periodo_id=periodo_id, laboratorio=lab, loja_id=loja_id_obs, texto=texto))
        db.session.commit()
        return jsonify({'mensagem': 'Salvo!'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'erro': str(e)}), 500


@app.route('/admin/preparar_envio/<int:periodo_id>')
def preparar_envio(periodo_id):
    if 'usuario_id' not in session or session.get('funcao') != 'Admin':
        return redirect(url_for('login'))

    periodo = (db.session.get(Periodo, periodo_id) or abort(404))
    lojas   = Loja.query.order_by(Loja.nome).all()

    # Fornecedores do grupo com contagem de itens
    laboratorios_do_grupo = db.session.query(Produto.laboratorio)\
        .filter_by(grupo=periodo.grupo_filtro).distinct().all()

    fornecedores_resumo = []
    for lab in laboratorios_do_grupo:
        nome_lab = lab[0]
        forn     = Fornecedor.query.filter_by(nome=nome_lab).first()
        email    = forn.email if forn else None
        qtd      = db.session.query(db.func.count(ItemPedido.id))\
            .join(Produto).join(Pedido)\
            .filter(
                Pedido.periodo_id  == periodo.id,
                Produto.laboratorio == nome_lab
            ).scalar()
        fornecedores_resumo.append({'nome': nome_lab, 'email': email, 'qtd_itens': qtd})

    # ── NOVO: todos os usuários verificados para a lista de confirmação ──
    # Inclui admins e gerentes de loja, cada um com sua loja (se houver)
    gerentes_disponiveis = Usuario.query.filter_by(verificado=True)\
        .order_by(Usuario.funcao, Usuario.nome).all()

    return render_template(
        'admin_envio_email.html',
        periodo=periodo,
        lojas=lojas,
        fornecedores_resumo=fornecedores_resumo,
        gerentes_disponiveis=gerentes_disponiveis,   # ← novo
    )


@csrf.exempt
@app.route('/admin/processar_envio_massa', methods=['POST'])
def processar_envio_massa():
    if 'usuario_id' not in session or session.get('funcao') != 'Admin':
        return jsonify({'erro': 'Não autorizado'}), 401

    admin_user      = db.session.get(Usuario, session['usuario_id'])
    remetente_email = admin_user.smtp_email if admin_user.smtp_email else admin_user.email
    remetente_senha = revelar_senha(admin_user.smtp_senha)
    smtp_server     = admin_user.smtp_server or 'smtp.locaweb.com.br'
    smtp_port       = admin_user.smtp_port   or 587

    if not remetente_senha:
        return jsonify({
            'erro': f'O usuário {admin_user.nome} não tem senha de e-mail configurada.'
        }), 400

    periodo_id          = request.form.get('periodo_id')
    periodo             = (db.session.get(Periodo, periodo_id) or abort(404))
    lojas_selecionadas_ids = request.form.getlist('lojas_ids')
    labs_selecionados   = request.form.getlist('labs_selecionados')
    lojas_filtradas     = [l for l in Loja.query.all() if str(l.id) in lojas_selecionadas_ids]

    # ── Destinatários de confirmação selecionados na tela ───────────────
    confirmacao_ids = request.form.getlist('confirmacao_ids')   # IDs de usuários
    emails_extras_raw = request.form.get('emails_extras', '').strip()

    # Busca e-mails dos usuários marcados
    emails_confirmacao = []
    if confirmacao_ids:
        usuarios_marcados = Usuario.query.filter(
            Usuario.id.in_([int(i) for i in confirmacao_ids]),
            Usuario.verificado == True
        ).all()
        emails_confirmacao = [u.email for u in usuarios_marcados if u.email]

    # Adiciona e-mails extras digitados manualmente
    if emails_extras_raw:
        extras = [e.strip() for e in emails_extras_raw.split(',') if e.strip()]
        emails_confirmacao += extras

    # Remove duplicatas mantendo ordem
    emails_confirmacao = list(dict.fromkeys(emails_confirmacao))

    # ── Mês e ano por extenso para o assunto ────────────────────────────
    MESES = {
        1:'Janeiro', 2:'Fevereiro', 3:'Março',    4:'Abril',
        5:'Maio',    6:'Junho',     7:'Julho',     8:'Agosto',
        9:'Setembro',10:'Outubro', 11:'Novembro', 12:'Dezembro'
    }
    mes_extenso = MESES.get(periodo.mes, str(periodo.mes))
    assunto_forn_base = f"Pedidos Rancho Alegre {mes_extenso}/{periodo.ano}"

    enviados_log = []
    erros_log    = []

    for lab_nome in labs_selecionados:
        forn = Fornecedor.query.filter_by(nome=lab_nome).first()

        if not (forn and forn.email):
            erros_log.append(f"{lab_nome}: Sem e-mail cadastrado.")
            continue

        try:
            excel_bytes = _excel_unico(
                periodo, lab_nome, lojas_filtradas, db, Produto, ItemPedido, Pedido
            )

            # ── E-MAIL 1: Fornecedor ─────────────────────────────────────
            assunto_forn = f"{assunto_forn_base} - {lab_nome}"
            corpo_forn   = (
                f"Prezado(a),\n\n"
                f"Segue em anexo os pedidos, conforme negociado.\n\n"
                f"Solicitamos, por gentileza, a confirmação do prazo de "
                f"faturamento e entrega.\n\n"
                f"Atenciosamente,\n"
                f"Setor de Compras\n"
                f"Lojas Rancho Alegre"
            )

            enviar_email_pedido(
                remetente_email=remetente_email,
                remetente_senha=remetente_senha,
                smtp_server=smtp_server,
                smtp_port=smtp_port,
                destinatario=forn.email,
                lista_cc=[remetente_email],
                assunto=assunto_forn,
                corpo=corpo_forn,
                anexo_bytes=excel_bytes,
                nome_anexo=f"Pedido_{lab_nome.replace(' ', '_')}.xlsx",
            )

            # ── E-MAIL 2: Confirmação para destinatários selecionados ────
            if emails_confirmacao:
                nomes_lojas  = ", ".join(l.nome for l in lojas_filtradas)
                corpo_confirm = (
                    f"Olá,\n\n"
                    f"O pedido do fornecedor {lab_nome} referente ao período "
                    f"{periodo.nome} foi enviado com sucesso.\n\n"
                    f"Lojas incluídas: {nomes_lojas}\n\n"
                    f"Segue em anexo a planilha enviada ao fornecedor para sua conferência.\n\n"
                    f"Atenciosamente,\n"
                    f"Setor de Compras\n"
                    f"Lojas Rancho Alegre"
                )
                for email_dest in emails_confirmacao:
                    try:
                        enviar_email_pedido(
                            remetente_email=remetente_email,
                            remetente_senha=remetente_senha,
                            smtp_server=smtp_server,
                            smtp_port=smtp_port,
                            destinatario=email_dest,
                            lista_cc=[remetente_email],
                            assunto=f"[Confirmação] {assunto_forn}",
                            corpo=corpo_confirm,
                            anexo_bytes=excel_bytes,
                            nome_anexo=f"Pedido_{lab_nome.replace(' ', '_')}.xlsx",
                        )
                    except Exception as eg:
                        app.logger.warning(
                            f"Falha ao notificar {email_dest} sobre {lab_nome}: {eg}"
                        )

            dest_confirmacao_str = ', '.join(emails_confirmacao) if emails_confirmacao else 'nenhum'
            enviados_log.append(
                f"{lab_nome} → fornecedor: {forn.email} | "
                f"confirmação: {dest_confirmacao_str}"
            )

            registrar_log(
                acao='EMAIL_ENVIADO',
                entidade='Fornecedor', entidade_id=forn.id,
                detalhe=(
                    f'Período {periodo_id} | Fornecedor: {forn.email} | '
                    f'Confirmação: {dest_confirmacao_str}'
                )
            )
            db.session.commit()
            app.logger.info(
                f"E-mails enviados para {lab_nome} "
                f"(fornecedor + {len(emails_confirmacao)} confirmação)."
            )

        except Exception as e:
            erros_log.append(f"Erro {lab_nome}: {str(e)}")
            app.logger.error(f"Falha ao enviar e-mail para {lab_nome}: {e}")

    return jsonify({'enviados': enviados_log, 'erros': erros_log})

@app.route('/admin/exportar_excel/<int:periodo_id>')
def exportar_excel(periodo_id):
    if 'usuario_id' not in session or session.get('funcao') != 'Admin':
        return redirect(url_for('login'))

    periodo = (db.session.get(Periodo, periodo_id) or abort(404))

    # Ordem personalizada das lojas
    _ordem = ["Pesqueira", "Recife", "Campina Grande", "Natal", "Maceió"]
    lojas  = sorted(Loja.query.all(), key=lambda l: _ordem.index(l.nome) if l.nome in _ordem else 999)

    produtos = Produto.query.filter_by(
        grupo=periodo.grupo_filtro
    ).order_by(Produto.laboratorio, Produto.nome).all()

    dados_lab = {}
    for p in produtos:
        if p.laboratorio not in dados_lab:
            dados_lab[p.laboratorio] = []
        negoc   = Negociacao.query.filter_by(periodo_id=periodo.id, produto_id=p.id).first()
        desconto = negoc.desconto    if negoc else 0.0
        bonif    = negoc.bonificacao if negoc else 0.0
        qtds_lojas = {}
        total_prod = 0
        for loja in lojas:
            qtd = db.session.query(db.func.sum(ItemPedido.quantidade)).join(Pedido).filter(
                ItemPedido.produto_id == p.id,
                Pedido.loja_id        == loja.id,
                Pedido.periodo_id     == periodo.id
            ).scalar() or 0
            qtds_lojas[loja.id] = qtd
            total_prod += qtd
        dados_lab[p.laboratorio].append({
            'nome': p.nome, 'caixa': p.unidade_caixa, 'preco': p.preco,
            'qtds': qtds_lojas, 'total': total_prod,
            'desconto': desconto, 'bonificacao': bonif
        })

    # FIX 11: geração delegada ao excel_service
    output       = gerar_excel_consolidado(periodo, lojas, dados_lab)
    nome_arquivo = f"Consolidado_{periodo.nome.replace('/', '-')}.xlsx"

    return send_file(
        output,
        download_name=nome_arquivo,
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/admin/auditoria')
def admin_auditoria():
    if 'usuario_id' not in session or session.get('funcao') != 'Admin':
        return redirect(url_for('login'))
    pagina = request.args.get('pagina', 1, type=int)
    filtro_acao = request.args.get('acao', '').strip()
    query = LogAuditoria.query
    if filtro_acao:
        query = query.filter(LogAuditoria.acao == filtro_acao)
    logs = query.order_by(LogAuditoria.data_hora.desc()).paginate(page=pagina, per_page=50, error_out=False)
    acoes_disponiveis = db.session.query(LogAuditoria.acao).distinct().order_by(LogAuditoria.acao).all()
    acoes_disponiveis = [a[0] for a in acoes_disponiveis]
    return render_template('admin_auditoria.html', logs=logs, filtro_acao=filtro_acao, acoes_disponiveis=acoes_disponiveis)

# --- SETUP (CRIAÇÃO DO BANCO) E INICIALIZAÇÃO ---
if __name__ == '__main__':
    with app.app_context():
        # Cria apenas tabelas que faltam sem apagar as outras
        db.create_all()
        
        # Garante Token Inicial
        if not Configuracao.query.filter_by(chave='admin_token').first():
            db.session.add(Configuracao(chave='admin_token', valor='123456'))
            db.session.commit()
            
        if not Loja.query.first():
            l1, l2, l3, l4, l5 = Loja(nome="Campina Grande"), Loja(nome="Maceió"), Loja(nome="Recife"), Loja(nome="Natal"), Loja(nome="Pesqueira")
            db.session.add_all([l1, l2, l3, l4, l5])
            
            # Usuário Admin Padrão
            if not Usuario.query.filter_by(email="admin@central.com").first():
                u2 = Usuario(nome="Comprador", email="admin@central.com", senha=generate_password_hash("1234"), funcao="Admin", verificado=True)
                db.session.add(u2)
                
            db.session.commit()

        # Configura Grupos Iniciais se a tabela estiver vazia
        if not Grupo.query.first():
            db.session.add(Grupo(nome="Medicamentos", codigo="2.1", dia_limite=10))
            db.session.add(Grupo(nome="Perfumaria", codigo="2.2", dia_limite=17))
            db.session.add(Grupo(nome="Diversos", codigo="DIVERSOS", dia_limite=1))
            db.session.commit()
    if os.name == 'nt': # 'nt' significa Windows
        print("Modo Desenvolvimento (Flask) iniciado...")
        app.run(host='0.0.0.0', port=5000, debug=True)
    else:
        print("Modo Produção (Waitress) iniciado...")
        serve(app, host='0.0.0.0', port=5000, threads=6)